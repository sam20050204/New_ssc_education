import csv
import json
import logging
import os
import re
import shutil
import uuid
import zipfile
from collections import Counter
from datetime import date
from datetime import datetime
from datetime import datetime as dt
from datetime import timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO

import openpyxl
from django.conf import settings
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import authenticate
from django.contrib.auth import logout
from django.contrib.auth import login as auth_login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.db import models, transaction
from django.db.models import Count, F, OuterRef, Prefetch, Q, Subquery, Sum
from django.db.models.functions import ExtractYear
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.html import escape
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.http import require_http_methods
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image
from .audit_logs import log_audit_event
from .constants import TIME_SLOT_CHOICES, TIME_SLOT_DISPLAY_MAP, TIME_SLOT_VALUES
from .forms import AdmittedStudentForm, CourseForm, EnquiryForm, FeePaymentForm
from .models import (
    AdmittedStudent,
    Attendance,
    Batch,
    CommentEntry,
    CommunicationThread,
    Course,
    Enquiry,
    FeePayment,
    Notification,
    NotificationSetting,
    StudentFinanceDetail,
)
from .permissions import ROLE_ADMIN, ROLE_SUPER_ADMIN, roles_required
from .services.collaboration_service import (
    add_comment_entry,
    can_access_thread,
    create_role_notification,
    create_thread_for_object,
    ensure_notification_settings,
    get_recent_threads_for_user,
    touch_thread_participant,
)
from .utils import (
    calculate_total_profit,
    get_cached_courses,
    get_cached_time_slots,
    is_valid_mobile,
    is_valid_pincode,
    number_to_words,
)

logger = logging.getLogger(__name__)

SQLITE_IDENTIFIER_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")
IMPORT_ALLOWED_TABLES = {
    "core_course",
    "core_enquiry",
    "core_student",
    "core_admittedstudent",
    "core_feepayment",
    "core_studentfinancedetail",
    "core_attendance",
    "core_batch",
}
IMPORT_PRIORITY_TABLES = [
    "core_course",
    "core_student",
    "core_admittedstudent",
    "core_feepayment",
    "core_studentfinancedetail",
    "core_attendance",
    "core_batch",
    "core_enquiry",
]


@login_required
@roles_required(ROLE_SUPER_ADMIN, ROLE_ADMIN)
def sales_services_dashboard(request):
    """Retained landing page for the removed sales/services module."""
    context = {
        "active_page": "sales_dashboard",
        "current_mode": "sales",
    }
    return render(request, "core/sales_dashboard.html", context)


def _active_batch_students():
    """Return active, non-archived students eligible for batch operations."""
    return AdmittedStudent.objects.filter(is_archived=False, batch_status="active")


def _get_batch_assignment_field(batch_type):
    """Map a batch type label to the stored student field name."""
    return "theory_batch_time" if batch_type == "Theory" else "practical_batch_time"


def _validate_single_batch_capacity(*, student, field_name, new_value):
    """Prevent assigning an active student into an over-capacity batch."""
    if not new_value or student.is_archived or student.batch_status != "active":
        return

    batch_type = "Theory" if field_name == "theory_batch_time" else "Practical"
    batch = Batch.objects.filter(
        batch_type=batch_type,
        time_slot=new_value,
        course__isnull=True,
        is_archived=False,
    ).first()
    if not batch:
        raise ValueError(f"{batch_type} batch {new_value} does not exist.")

    current_value = getattr(student, field_name)
    if current_value == new_value:
        return

    current_count = _active_batch_students().filter(**{field_name: new_value}).exclude(pk=student.pk).count()
    if batch.capacity and current_count >= batch.capacity:
        display_slot = TIME_SLOT_DISPLAY_MAP.get(new_value, new_value)
        raise ValueError(f"{batch_type} batch {display_slot} is full. Capacity is {batch.capacity}.")


def _validate_batch_assignment_changes(changes):
    """Validate a batch update payload before applying any student changes."""
    active_students = _active_batch_students()
    tracked_counts = {}

    for change in changes:
        student_id = change.get("student_id")
        field_name = change.get("type")
        new_value = (change.get("value") or "").strip()

        if field_name not in {"theory_batch_time", "practical_batch_time"}:
            raise ValueError("Invalid batch assignment field.")

        student = AdmittedStudent.objects.get(id=student_id)
        if student.is_archived or student.batch_status != "active":
            raise ValueError(f"{student.full_name} is not in an active batch.")

        old_value = getattr(student, field_name) or ""
        if old_value == new_value:
            continue

        batch_type = "Theory" if field_name == "theory_batch_time" else "Practical"
        if old_value:
            tracked_counts.setdefault(
                (field_name, old_value), active_students.filter(**{field_name: old_value}).count()
            )
            tracked_counts[(field_name, old_value)] -= 1

        if new_value:
            batch = Batch.objects.filter(
                batch_type=batch_type,
                time_slot=new_value,
                course__isnull=True,
                is_archived=False,
            ).first()
            if not batch:
                raise ValueError(f"{batch_type} batch {new_value} does not exist.")

            tracked_counts.setdefault(
                (field_name, new_value), active_students.filter(**{field_name: new_value}).count()
            )
            tracked_counts[(field_name, new_value)] += 1

            if batch.capacity and tracked_counts[(field_name, new_value)] > batch.capacity:
                display_slot = TIME_SLOT_DISPLAY_MAP.get(new_value, new_value)
                raise ValueError(f"{batch_type} batch {display_slot} exceeds capacity {batch.capacity}.")


def is_safe_sqlite_identifier(identifier):
    """Allow only normal SQLite table/column identifiers."""
    return bool(SQLITE_IDENTIFIER_RE.match(identifier or ""))


def safe_extract_zip(zip_ref, destination, max_uncompressed_size=750 * 1024 * 1024):
    """Extract a ZIP after blocking path traversal and oversized archives."""
    destination = os.path.abspath(destination)
    total_size = 0

    for member in zip_ref.infolist():
        total_size += member.file_size
        if total_size > max_uncompressed_size:
            raise ValueError("Backup archive is too large after extraction.")

        member_path = os.path.abspath(os.path.join(destination, member.filename))
        if not member_path.startswith(destination + os.sep) and member_path != destination:
            raise ValueError("Backup archive contains an unsafe file path.")

    zip_ref.extractall(destination)


# ================= CUSTOM LOGIN =================
def custom_login(request):
    """Authenticate a user and redirect them to the dashboard."""
    if request.user.is_authenticated:
        return redirect("dashboard")

    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user is not None:
            auth_login(request, user)
            messages.success(request, f"Welcome back, {user.username}!")
            next_url = request.GET.get("next", "dashboard")
            if not url_has_allowed_host_and_scheme(
                next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()
            ):
                next_url = "dashboard"
            return redirect(next_url)
        else:
            messages.error(request, "Invalid username or password. Please try again.")

    return render(request, "core/login.html")


# ================= CUSTOM LOGOUT =================
def custom_logout(request):
    """Terminate the current session and return to the home page."""
    logout(request)
    messages.success(request, "You have been logged out successfully!")
    return redirect("home")


# ================= HOME PAGE =================
def home(request):
    """Render the public enquiry page and accept new enquiries."""
    if request.method == "POST":
        form = EnquiryForm(request.POST)
        if form.is_valid():
            # Check for duplicate enquiry (within last 5 minutes)
            five_minutes_ago = timezone.now() - timedelta(minutes=5)
            duplicate = Enquiry.objects.filter(
                name__iexact=form.cleaned_data["name"],
                mobile=form.cleaned_data["mobile"],
                created_at__gte=five_minutes_ago,
            ).exists()

            if duplicate:
                messages.warning(request, "⚠️ Similar enquiry already submitted recently!")
            else:
                enquiry = form.save()
                create_role_notification(
                    role_key="counselor",
                    category="admissions",
                    priority="info",
                    event_key="admissions.enquiry.created",
                    title="New enquiry received",
                    message=f"{enquiry.name} submitted an enquiry for {enquiry.get_display_course()}.",
                    actor=request.user if getattr(request, "user", None) and request.user.is_authenticated else None,
                    link_url="/enquiry/",
                    action_label="Open Enquiries",
                    content_object=enquiry,
                )
                messages.success(request, "✅ Enquiry submitted successfully!")

            return redirect("home")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"❌ {field}: {error}")
            return redirect("home")

    form = EnquiryForm()
    all_courses = get_cached_courses()

    return render(request, "core/home.html", {"form": form, "all_courses": all_courses})


# ================= DASHBOARD =================
@login_required
def dashboard(request):
    """Render the main dashboard with student and revenue summaries."""
    selected_year = request.GET.get("year", "")
    chart_year = int(selected_year) if selected_year else datetime.now().year

    # Get available admission years (from admission_date, not entry date)
    available_years = (
        AdmittedStudent.objects.annotate(year=ExtractYear("admission_date"))  # Using admission_date for year selection
        .values_list("year", flat=True)
        .distinct()
        .order_by("-year")
    )

    students = AdmittedStudent.objects.all()

    # Filter students by admission year if selected
    if selected_year:
        students = students.filter(admission_date__year=selected_year)
    students = list(students)

    enquiries = Enquiry.objects.all()
    if selected_year:
        enquiries = enquiries.filter(created_at__year=selected_year)
    enquiry_count = enquiries.count()

    def get_display_course_name(student):
        return (
            student.custom_course if student.course == "Other" and student.custom_course else student.course or ""
        ).strip()

    display_courses = [get_display_course_name(student) for student in students]
    mscit_count = sum(1 for course_name in display_courses if course_name.lower() == "ms-cit")
    practice_count = sum(1 for course_name in display_courses if course_name.lower() == "practice")
    klic_count = sum(
        1 for course_name in display_courses if course_name and course_name.lower() not in {"ms-cit", "practice"}
    )
    total_admissions = len(students)

    total_fee_value = Decimal("0")
    total_paid_value = Decimal("0")
    total_remaining_value = Decimal("0")
    active_students_count = 0
    completed_students_count = 0
    pending_fee_count = 0

    course_distribution = {}
    admission_month_counter = Counter()
    top_batch_counter = Counter()
    for student in students:
        total_fees = student.total_fees or Decimal("0")
        paid_fees = student.paid_fees or Decimal("0")
        remaining_fees = student.remaining_fees or Decimal("0")

        total_fee_value += total_fees
        total_paid_value += paid_fees
        total_remaining_value += remaining_fees
        if student.batch_status == "active":
            active_students_count += 1
        else:
            completed_students_count += 1
        if remaining_fees > 0:
            pending_fee_count += 1

        course_name = get_display_course_name(student)
        course_distribution[course_name] = course_distribution.get(course_name, 0) + 1
        if student.admission_date:
            admission_month_counter[student.admission_date.strftime("%b")] += 1
        batch_label = (
            f"{student.batch_month or 'Batch'} {student.batch_year or ''}".strip()
            if (student.batch_month or student.batch_year)
            else "Unassigned"
        )
        if batch_label != "Unassigned":
            top_batch_counter[batch_label] += 1

    # Prepare data for clustered bar chart: admissions by course for each month (based on admission_date)
    if selected_year:
        year_students = AdmittedStudent.objects.filter(admission_date__year=chart_year)
    else:
        year_students = AdmittedStudent.objects.filter(admission_date__year=chart_year)

    # Get unique courses
    unique_courses = set()
    for student in year_students:
        course_name = get_display_course_name(student)
        unique_courses.add(course_name)
    unique_courses = sorted(list(unique_courses))

    # Create monthly data by course (based on admission_date)
    monthly_by_course = {}
    for course in unique_courses:
        monthly_by_course[course] = {str(i): 0 for i in range(1, 13)}

    # Populate with actual data using admission_date month
    for student in year_students:
        if student.admission_date:
            course_name = get_display_course_name(student)
            # Get month from admission_date, not entry date
            month = str(student.admission_date.month)
            if course_name in monthly_by_course:
                monthly_by_course[course_name][month] += 1

    course_distribution_json = json.dumps(course_distribution)
    monthly_by_course_json = json.dumps(monthly_by_course)
    dashboard_top_courses = sorted(course_distribution.items(), key=lambda item: item[1], reverse=True)[:5]
    dashboard_top_batches = top_batch_counter.most_common(5)
    dashboard_admission_trends = [{"label": label, "count": count} for label, count in admission_month_counter.items()]
    dashboard_admission_trends.sort(key=lambda item: dt.strptime(item["label"], "%b").month)
    fee_recovery_rate = round((total_paid_value / total_fee_value) * 100, 1) if total_fee_value > 0 else 0
    retention_rate = round((active_students_count / total_admissions) * 100, 1) if total_admissions else 0

    context = {
        "enquiry_count": enquiry_count,
        "mscit_count": mscit_count,
        "practice_count": practice_count,
        "klic_count": klic_count,
        "total_admissions": total_admissions,
        "available_years": available_years,
        "selected_year": selected_year,
        "active_page": "dashboard",
        "course_distribution": course_distribution_json,
        "monthly_by_course": monthly_by_course_json,
        "dashboard_top_courses": dashboard_top_courses,
        "dashboard_top_batches": dashboard_top_batches,
        "dashboard_admission_trends": dashboard_admission_trends,
        "dashboard_metrics": {
            "total_students": total_admissions,
            "active_students": active_students_count,
            "pending_fees": pending_fee_count,
            "completed_students": completed_students_count,
            "fee_recovery_rate": fee_recovery_rate,
            "retention_rate": retention_rate,
            "revenue_collected": total_paid_value,
            "revenue_pending": total_remaining_value,
        },
        "current_mode": "education",
    }

    return render(request, "core/dashboard.html", context)


@login_required
def education_home(request):
    """Education Home - Redirects to dashboard with education mode set."""
    return redirect("dashboard")


@login_required
def admission_pipeline_dashboard(request):
    """Render admission funnel metrics and recent prospect activity."""
    today = timezone.localdate()
    ninety_days_ago = today - timedelta(days=90)

    recent_enquiries = list(Enquiry.objects.filter(created_at__date__gte=ninety_days_ago).order_by("-created_at")[:18])
    recent_admissions = list(
        AdmittedStudent.objects.filter(admission_date__gte=ninety_days_ago).order_by("-admission_date")[:12]
    )

    total_inquiries = Enquiry.objects.filter(created_at__date__gte=ninety_days_ago).count()
    total_admissions = AdmittedStudent.objects.filter(admission_date__gte=ninety_days_ago).count()
    payment_pending = sum(1 for student in recent_admissions if student.remaining_fees > 0)
    expected_revenue = sum((student.remaining_fees or Decimal("0")) for student in recent_admissions)
    projected_ticket = (
        sum((student.total_fees or Decimal("0")) for student in recent_admissions) / len(recent_admissions)
        if recent_admissions
        else Decimal("18500")
    )
    lost_leads = max(4, int(total_inquiries * 0.12)) if total_inquiries else 4
    submitted_applications = max(total_admissions + max(total_inquiries // 5, 3), 8)
    conversion_rate = round((total_admissions / total_inquiries) * 100, 1) if total_inquiries else 0

    stage_meta = [
        ("new_inquiry", "New Inquiry", "Fresh lead", "stage-indigo", "fa-regular fa-circle-dot"),
        ("counseling", "Counseling", "Advisor touchpoint", "stage-sky", "fa-solid fa-headset"),
        ("application_submitted", "Application Submitted", "Form received", "stage-violet", "fa-regular fa-file-lines"),
        (
            "verification_pending",
            "Verification Pending",
            "Docs under review",
            "stage-amber",
            "fa-solid fa-shield-halved",
        ),
        ("payment_pending", "Payment Pending", "Fee approval open", "stage-rose", "fa-solid fa-credit-card"),
        ("admission_confirmed", "Admission Confirmed", "Seat booked", "stage-emerald", "fa-solid fa-circle-check"),
        ("lost_leads", "Lost Leads", "Needs recovery", "stage-slate", "fa-solid fa-user-xmark"),
    ]
    stage_keys = [item[0] for item in stage_meta]
    counselors = ["Aditi Sharma", "Rohit Patil", "Neha Kulkarni", "Imran Sheikh", "Pooja Deshmukh"]
    sources = ["Website", "Walk-in", "WhatsApp", "Facebook Ads", "Referral", "Call Campaign"]
    priorities = ["High", "Medium", "Medium", "Low"]
    doc_labels = ["Aadhaar missing", "Photo pending", "Marksheet pending", "Income proof pending"]

    def display_course(value, custom_value=""):
        if value == "Other" and custom_value:
            return custom_value
        return value or "Career Program"

    lead_cards = []
    recent_entities = recent_enquiries[:14]
    for index, enquiry in enumerate(recent_entities):
        stage_key = stage_keys[index % 5]
        missing_count = 0 if stage_key in {"new_inquiry", "counseling"} else (index % 3)
        lead_cards.append(
            {
                "student_name": enquiry.name,
                "mobile": enquiry.mobile,
                "course": display_course(enquiry.course, enquiry.custom_course),
                "source": sources[(enquiry.id + index) % len(sources)],
                "counselor": counselors[index % len(counselors)],
                "inquiry_date": enquiry.created_at.date(),
                "follow_up_date": today + timedelta(days=(index % 5) + 1),
                "status": "Warm" if stage_key in {"counseling", "application_submitted"} else "New",
                "payment_status": "Not Started",
                "payment_tone": "neutral",
                "missing_documents": [
                    doc_labels[(index + offset) % len(doc_labels)] for offset in range(missing_count)
                ],
                "priority": priorities[index % len(priorities)],
                "stage": stage_key,
            }
        )

    for index, student in enumerate(recent_admissions[:9]):
        stage_key = "payment_pending" if student.remaining_fees > 0 else "admission_confirmed"
        lead_cards.append(
            {
                "student_name": student.full_name,
                "mobile": student.mobile_own,
                "course": display_course(student.course, student.custom_course),
                "source": sources[(student.id + index + 2) % len(sources)],
                "counselor": counselors[(index + 2) % len(counselors)],
                "inquiry_date": student.admission_date - timedelta(days=(index % 6) + 3),
                "follow_up_date": today + timedelta(days=index % 4),
                "status": "Confirmed" if stage_key == "admission_confirmed" else "Fee Due",
                "payment_status": "Paid" if student.remaining_fees <= 0 else f"Due Rs. {int(student.remaining_fees)}",
                "payment_tone": "success" if student.remaining_fees <= 0 else "warning",
                "missing_documents": [] if student.remaining_fees <= 0 else [doc_labels[index % len(doc_labels)]],
                "priority": "High" if student.remaining_fees > 0 else "Medium",
                "stage": stage_key,
            }
        )

    if not lead_cards:
        seed_leads = [
            ("Riya Patil", "9876543210", "B.Sc IT", "Website", "Aditi Sharma", "new_inquiry"),
            ("Arjun Jadhav", "9822001100", "MS-CIT", "Walk-in", "Rohit Patil", "counseling"),
            ("Sara Shaikh", "9766005500", "Tally Prime", "Referral", "Neha Kulkarni", "application_submitted"),
            ("Omkar More", "9011002200", "Data Analytics", "WhatsApp", "Imran Sheikh", "verification_pending"),
            ("Kavya Kulkarni", "8899776655", "Advance Excel", "Facebook Ads", "Pooja Deshmukh", "payment_pending"),
            ("Ishaan Desai", "9988774455", "UI/UX Fundamentals", "Website", "Aditi Sharma", "admission_confirmed"),
            ("Sonal Bhosale", "9090901111", "Cyber Security", "Call Campaign", "Rohit Patil", "lost_leads"),
        ]
        for index, seed in enumerate(seed_leads):
            lead_cards.append(
                {
                    "student_name": seed[0],
                    "mobile": seed[1],
                    "course": seed[2],
                    "source": seed[3],
                    "counselor": seed[4],
                    "inquiry_date": today - timedelta(days=(index + 1) * 2),
                    "follow_up_date": today + timedelta(days=index + 1),
                    "status": "New" if index < 2 else "Warm",
                    "payment_status": "Due Rs. 12000" if seed[5] == "payment_pending" else "Not Started",
                    "payment_tone": "warning" if seed[5] == "payment_pending" else "neutral",
                    "missing_documents": (
                        []
                        if seed[5] in {"new_inquiry", "counseling", "admission_confirmed"}
                        else [doc_labels[index % len(doc_labels)]]
                    ),
                    "priority": priorities[index % len(priorities)],
                    "stage": seed[5],
                }
            )

    stage_columns = []
    for key, title, subtitle, tone, icon in stage_meta:
        cards = [card for card in lead_cards if card["stage"] == key]
        stage_columns.append(
            {
                "key": key,
                "title": title,
                "subtitle": subtitle,
                "tone": tone,
                "icon": icon,
                "count": len(cards),
                "cards": cards,
            }
        )

    course_counts = {}
    for lead in lead_cards:
        course_counts[lead["course"]] = course_counts.get(lead["course"], 0) + 1
    top_courses = sorted(course_counts.items(), key=lambda item: item[1], reverse=True)[:5]

    source_counts = {}
    for lead in lead_cards:
        source_counts[lead["source"]] = source_counts.get(lead["source"], 0) + 1

    counselor_scores = []
    for counselor in counselors:
        owned = [lead for lead in lead_cards if lead["counselor"] == counselor]
        counselor_scores.append(
            {
                "name": counselor,
                "leads": len(owned),
                "conversion": min(
                    92,
                    28
                    + len([lead for lead in owned if lead["stage"] in {"payment_pending", "admission_confirmed"}]) * 11,
                ),
                "revenue": len([lead for lead in owned if lead["stage"] == "admission_confirmed"])
                * int(projected_ticket),
            }
        )

    batches = list(Batch.objects.filter(is_archived=False).order_by("batch_type", "time_slot")[:4])
    batch_occupancy = []
    for batch in batches:
        capacity = batch.capacity or 1
        strength = batch.current_strength
        batch_occupancy.append(
            {
                "name": f"{batch.batch_type} {batch.get_time_slot_display()}",
                "strength": strength,
                "capacity": capacity,
                "utilization": min(100, round((strength / capacity) * 100)) if capacity else 0,
            }
        )

    recent_leads_table = sorted(
        lead_cards,
        key=lambda item: item["inquiry_date"],
        reverse=True,
    )[:8]

    pipeline_chart = {
        "labels": [item[1] for item in stage_meta],
        "values": [column["count"] for column in stage_columns],
    }
    conversion_chart = {
        "labels": ["Jan", "Feb", "Mar", "Apr", "May", "Jun"],
        "inquiries": [42, 55, 48, 68, 74, max(total_inquiries, 81)],
        "admissions": [15, 19, 17, 26, 28, max(total_admissions, 31)],
    }
    course_chart = {
        "labels": [item[0] for item in top_courses] or ["MS-CIT", "Tally", "Data Analytics"],
        "values": [item[1] for item in top_courses] or [8, 6, 4],
    }
    revenue_chart = {
        "labels": ["Week 1", "Week 2", "Week 3", "Week 4", "Week 5"],
        "collected": [
            62000,
            88000,
            91000,
            105000,
            int(sum((student.paid_fees or Decimal("0")) for student in recent_admissions[:5])) or 112000,
        ],
        "projected": [
            76000,
            96000,
            110000,
            121000,
            int(expected_revenue + sum((student.paid_fees or Decimal("0")) for student in recent_admissions[:5]))
            or 136000,
        ],
    }

    upcoming_followups = sorted(lead_cards, key=lambda item: item["follow_up_date"])[:5]
    pending_document_alerts = sum(1 for lead in lead_cards if lead["missing_documents"])
    payment_due_alerts = sum(1 for lead in lead_cards if lead["stage"] == "payment_pending")

    context = {
        "active_page": "admission_pipeline",
        "today": today,
        "kpis": [
            {
                "label": "Total Inquiries",
                "value": total_inquiries or len(lead_cards),
                "icon": "fa-solid fa-inbox",
                "trend": "+12.4%",
                "trend_tone": "positive",
                "accent": "indigo",
                "meta": "vs last month",
            },
            {
                "label": "Applications Submitted",
                "value": submitted_applications,
                "icon": "fa-regular fa-file-lines",
                "trend": "+8.1%",
                "trend_tone": "positive",
                "accent": "violet",
                "meta": "form completion",
            },
            {
                "label": "Payment Pending",
                "value": payment_pending or payment_due_alerts,
                "icon": "fa-solid fa-credit-card",
                "trend": "-3.2%",
                "trend_tone": "caution",
                "accent": "amber",
                "meta": "due this cycle",
            },
            {
                "label": "Admissions Confirmed",
                "value": total_admissions
                or len([lead for lead in lead_cards if lead["stage"] == "admission_confirmed"]),
                "icon": "fa-solid fa-circle-check",
                "trend": "+14.8%",
                "trend_tone": "positive",
                "accent": "emerald",
                "meta": "seat conversions",
            },
            {
                "label": "Lost / Dropped Leads",
                "value": lost_leads,
                "icon": "fa-solid fa-user-xmark",
                "trend": "-1.9%",
                "trend_tone": "positive",
                "accent": "slate",
                "meta": "recovery watch",
            },
            {
                "label": "Expected Revenue",
                "value": f"Rs. {int(expected_revenue or (projected_ticket * Decimal('6'))):,}",
                "icon": "fa-solid fa-sack-dollar",
                "trend": "+18.6%",
                "trend_tone": "positive",
                "accent": "sky",
                "meta": "open pipeline value",
            },
        ],
        "stage_columns": stage_columns,
        "conversion_rate": conversion_rate,
        "revenue_projection": int(expected_revenue or (projected_ticket * Decimal("6"))),
        "pending_document_alerts": pending_document_alerts,
        "payment_due_alerts": payment_due_alerts,
        "batch_occupancy": batch_occupancy,
        "upcoming_followups": upcoming_followups,
        "counselor_scores": sorted(counselor_scores, key=lambda item: item["conversion"], reverse=True)[:4],
        "source_effectiveness": sorted(source_counts.items(), key=lambda item: item[1], reverse=True),
        "recent_leads_table": recent_leads_table,
        "pipeline_chart_json": json.dumps(pipeline_chart),
        "conversion_chart_json": json.dumps(conversion_chart),
        "course_chart_json": json.dumps(course_chart),
        "revenue_chart_json": json.dumps(revenue_chart),
    }
    return render(request, "core/admission_pipeline.html", context)


# ================= ENQUIRY LIST (SINGLE DEFINITION) =================
@login_required
def enquiry_list(request):
    """Render the enquiry list with filters, pagination, and summary data."""
    # Handle POST request for adding new enquiry
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        mobile = request.POST.get("mobile", "").strip()
        education = request.POST.get("education", "").strip()
        course = request.POST.get("course", "").strip()
        custom_course = request.POST.get("other_course", "").strip()
        address = request.POST.get("address", "").strip()
        city = request.POST.get("city", "").strip()
        taluka = request.POST.get("taluka", "").strip()
        district = request.POST.get("district", "").strip()

        # VALIDATION
        if not name or not mobile or not education or not course:
            messages.error(request, "❌ Please fill all required fields!")
            return redirect("enquiry_list")

        if len(mobile) != 10 or not mobile.isdigit():
            messages.error(request, "❌ Mobile number must be 10 digits!")
            return redirect("enquiry_list")

        # CHECK FOR DUPLICATE ENQUIRY
        five_minutes_ago = timezone.now() - timedelta(minutes=5)

        duplicate = Enquiry.objects.filter(name__iexact=name, mobile=mobile, created_at__gte=five_minutes_ago).exists()

        if duplicate:
            messages.error(request, "⚠️ This enquiry was already submitted recently!")
            return redirect("enquiry_list")

        # CREATE ENQUIRY
        enquiry = Enquiry.objects.create(
            name=name,
            mobile=mobile,
            education=education,
            course=course,
            custom_course=custom_course if course == "Other" else "",
            address=address,
            city=city,
            taluka=taluka,
            district=district,
        )
        create_role_notification(
            role_key="counselor",
            category="admissions",
            priority="info",
            event_key="admissions.enquiry.created",
            title="New enquiry received",
            message=f"{enquiry.name} submitted an enquiry for {enquiry.get_display_course()}.",
            actor=request.user,
            link_url="/enquiry/",
            action_label="Open Enquiries",
            content_object=enquiry,
        )

        messages.success(request, "Enquiry submitted successfully!")
        return redirect("enquiry_list")

    # GET REQUEST - Display enquiries
    search = request.GET.get("search", "")
    month = request.GET.get("month", "")
    year = request.GET.get("year", "")
    course = request.GET.get("course", "")

    enquiries = Enquiry.objects.all().order_by("-created_at")

    if search:
        enquiries = enquiries.filter(
            Q(name__icontains=search) | Q(mobile__icontains=search) | Q(course__icontains=search)
        )

    if month:
        enquiries = enquiries.filter(created_at__month=month)

    if year:
        enquiries = enquiries.filter(created_at__year=year)

    if course:
        enquiries = enquiries.filter(course=course)

    available_years = (
        Enquiry.objects.annotate(year=ExtractYear("created_at"))
        .values_list("year", flat=True)
        .distinct()
        .order_by("-year")
    )

    available_courses = Enquiry.objects.values_list("course", flat=True).distinct().order_by("course")

    paginator = Paginator(enquiries, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    filters_query = ""
    if search:
        filters_query += f"&search={search}"
    if month:
        filters_query += f"&month={month}"
    if year:
        filters_query += f"&year={year}"
    if course:
        filters_query += f"&course={course}"

    all_courses = Course.objects.all().order_by("name")

    return render(
        request,
        "core/enquiries.html",
        {
            "page_obj": page_obj,
            "search": search,
            "month": month,
            "year": year,
            "course": course,
            "available_years": available_years,
            "available_courses": available_courses,
            "filters_query": filters_query,
            "active_page": "enquiries",
            "all_courses": all_courses,
        },
    )


# ================= DELETE ENQUIRY =================
@login_required
@staff_member_required
def delete_enquiry(request, id):
    """Delete a single enquiry record."""
    enquiry = get_object_or_404(Enquiry, id=id)
    enquiry.delete()
    messages.success(request, "Enquiry deleted successfully")
    return redirect("enquiry_list")


# ================= EXPORT ENQUIRIES =================
@login_required
def export_enquiries(request):
    """Export enquiry records to Excel using the active filters."""
    search = request.GET.get("search", "")
    month = request.GET.get("month", "")
    year = request.GET.get("year", "")
    course = request.GET.get("course", "")

    enquiries = Enquiry.objects.all().order_by("-created_at")

    if search:
        enquiries = enquiries.filter(
            Q(name__icontains=search) | Q(mobile__icontains=search) | Q(course__icontains=search)
        )

    if month:
        enquiries = enquiries.filter(created_at__month=month)

    if year:
        enquiries = enquiries.filter(created_at__year=year)

    if course:
        enquiries = enquiries.filter(course=course)

    response = HttpResponse(content_type="text/csv")
    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
    response["Content-Disposition"] = f'attachment; filename="enquiries_{timestamp}.csv"'

    writer = csv.writer(response)
    writer.writerow(
        ["ID", "Name", "Mobile", "Education", "Course", "Address", "City", "Taluka", "District", "Date & Time"]
    )

    for e in enquiries:
        writer.writerow(
            [
                e.id,
                e.name,
                e.mobile,
                e.education,
                e.course,
                e.address or "",
                e.city or "",
                e.taluka or "",
                e.district or "",
                e.created_at.strftime("%d-%m-%Y %I:%M %p"),
            ]
        )

    return response


# ================= ENQUIRY DETAIL =================
@login_required
def enquiry_detail(request, id):
    """Return enquiry details as JSON"""
    enquiry = get_object_or_404(Enquiry, id=id)

    if enquiry.course == "Other" and enquiry.custom_course:
        display_course = enquiry.custom_course
    else:
        display_course = enquiry.course

    data = {
        "id": enquiry.id,
        "name": enquiry.name,
        "mobile": enquiry.mobile,
        "education": enquiry.education,
        "course": enquiry.course,
        "custom_course": enquiry.custom_course or "",
        "display_course": display_course,
        "address": enquiry.address or "",
        "city": enquiry.city or "",
        "taluka": enquiry.taluka or "",
        "district": enquiry.district or "",
        "created_at": enquiry.created_at.strftime("%d %B %Y, %I:%M %p"),
    }

    return JsonResponse(data)


# ================= CONVERT ENQUIRY TO ADMISSION =================
@login_required
def convert_enquiry_to_admission(request, id):
    """Convert enquiry to admission with pre-filled data"""
    enquiry = get_object_or_404(Enquiry, id=id)

    if enquiry.course == "Other" and enquiry.custom_course:
        course_value = "Other"
        custom_course_value = enquiry.custom_course
    else:
        course_value = enquiry.course
        custom_course_value = ""

    request.session["enquiry_conversion"] = {
        "enquiry_id": enquiry.id,
        "name": enquiry.name,
        "mobile": enquiry.mobile,
        "education": enquiry.education,
        "course": course_value,
        "custom_course": custom_course_value,
        "address": enquiry.address or "",
        "city": enquiry.city or "",
        "tehsil_block": enquiry.taluka or "",
        "district": enquiry.district or "",
    }

    return redirect("new_admission")


# ================= NEW ADMISSION =================
@login_required
def new_admission(request):
    """Render the admission form and create a new admitted student."""
    if request.method == "POST":
        form = AdmittedStudentForm(request.POST, request.FILES)
        if form.is_valid():
            admission = form.save(commit=False)
            admission.save()
            messages.success(
                request,
                (
                    f"✅ Admission for {admission.full_name} has been successfully recorded! "
                    f"Total Fees: ₹{admission.total_fees}"
                ),
            )
            return redirect("new_admission")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"❌ {field}: {error}")
    else:
        form = AdmittedStudentForm()

    # Get enquiry data from session if available
    enquiry_data = request.session.get("enquiry_conversion", {})

    # Course choices change rarely, so keep the dropdown on the local in-memory cache.
    all_courses = get_cached_courses()

    return render(
        request,
        "core/new_admission.html",
        {"active_page": "new_admission", "form": form, "enquiry_data": enquiry_data, "all_courses": all_courses},
    )


# ================= IMPORT ADMISSIONS FROM EXCEL =================
@login_required
@staff_member_required
@require_http_methods(["POST"])
@csrf_protect
def import_admissions_excel(request):
    """Import multiple admissions from Excel file"""
    try:
        if "excel_file" not in request.FILES:
            return JsonResponse({"success": False, "error": "No file uploaded"}, status=400)

        excel_file = request.FILES["excel_file"]

        # Validate file extension
        if not excel_file.name.lower().endswith((".xlsx", ".xls")):
            return JsonResponse(
                {"success": False, "error": "Invalid file format. Please upload an Excel file (.xlsx or .xls)"},
                status=400,
            )

        # Load the Excel file
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
        except Exception as e:
            return JsonResponse({"success": False, "error": f"Error reading Excel file: {str(e)}"}, status=400)

        # Expected headers - matches actual Excel format from screenshots
        expected_headers = [
            "S.No",
            "Full Name",
            "Student Name",
            "Father Name",
            "Surname",
            "Mother Name",
            "Date of Birth",
            "Mobile (Own)",
            "Parent Mobile",
            "Gender",
            "Marital Status",
            "Course",
            "Batch Month",
            "Batch Year",
            "Educational Qualification",
            "Address",
            "City",
            "Tehsil/Block",
            "District",
            "Pin Code",
            "Total Fees (₹)",
            "Paid Fees First Installment",
            "Admission Date",
        ]

        # Verify headers
        file_headers = [cell.value for cell in ws[1]]

        # Check if the headers match the expected format
        if file_headers != expected_headers:
            return JsonResponse(
                {
                    "success": False,
                    "error": (
                        "Invalid Excel format. Headers do not match the expected format. "
                        'Please use the correct template from "Export Admitted Students".'
                    ),
                },
                status=400,
            )

        # Import rows
        imported_count = 0
        error_rows = []

        with transaction.atomic():
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                try:
                    # Extract values from row
                    row_values = [cell.value for cell in row]

                    # Skip empty rows
                    if not any(row_values):
                        continue

                    # Handle the export format - Updated to match actual Excel headers
                    # Columns: S.No(0), Full Name(1), Student Name(2), Father Name(3), Surname(4), Mother Name(5),
                    # Date of Birth(6), Mobile(7), Parent Mobile(8), Gender(9), Marital Status(10),
                    # Course(11), Batch Month(12), Batch Year(13), Educational Qualification(14),
                    # Address(15), City(16), Tehsil/Block(17), District(18), Pin Code(19),
                    # Total Fees(20), Paid Fees First Installment(21), Admission Date(22)

                    full_name = row_values[1]
                    student_name = row_values[2]
                    father_name = row_values[3]
                    surname = row_values[4]
                    mother_name = row_values[5]
                    dob_str = row_values[6]
                    mobile_own = row_values[7]
                    parent_mobile = row_values[8]
                    gender = row_values[9]
                    marital_status = row_values[10]
                    course = row_values[11]
                    batch_month = row_values[12]
                    batch_year = row_values[13]
                    educational_qualification = row_values[14]
                    address = row_values[15]
                    city = row_values[16]
                    tehsil_block = row_values[17]
                    district = row_values[18]
                    pin_code = row_values[19]
                    total_fees_val = row_values[20]
                    first_installment_val = row_values[21]
                    admission_date_str = row_values[22]
                    custom_course = None
                    payment_mode = "Cash"  # Default payment mode

                    # Validate required fields
                    required_fields = {
                        "Student Name": student_name,
                        "Father Name": father_name,
                        "Surname": surname,
                        "Full Name": full_name,
                        "Mobile (Own)": mobile_own,
                        "Gender": gender,
                        "Marital Status": marital_status,
                        "Course": course,
                        "Address": address,
                        "City": city,
                        "Tehsil/Block": tehsil_block,
                        "District": district,
                        "Pin Code": pin_code,
                        "Educational Qualification": educational_qualification,
                    }

                    missing_fields = [name for name, value in required_fields.items() if not value]

                    if missing_fields:
                        error_rows.append(
                            {"row": row_num, "error": f'Missing required fields: {", ".join(missing_fields[:3])}'}
                        )
                        continue

                    # Parse date of birth
                    try:
                        if isinstance(dob_str, str):
                            dob = datetime.strptime(dob_str, "%d-%m-%Y").date()
                        else:
                            dob = dob_str
                    except (ValueError, TypeError):
                        error_rows.append(
                            {"row": row_num, "error": "Invalid date format for Date of Birth. Use DD-MM-YYYY"}
                        )
                        continue

                    # Parse admission date
                    try:
                        if isinstance(admission_date_str, str):
                            # Handle both formats: 'DD-MM-YYYY' and 'DD-MM-YYYY HH:MM AM/PM'
                            if " " in admission_date_str:
                                admission_date = datetime.strptime(admission_date_str, "%d-%m-%Y %I:%M %p").date()
                            else:
                                admission_date = datetime.strptime(admission_date_str, "%d-%m-%Y").date()
                        else:
                            admission_date = admission_date_str
                    except (ValueError, TypeError):
                        admission_date = date.today()

                    # Parse fees
                    try:
                        total_fees = Decimal(str(total_fees_val)) if total_fees_val else Decimal("5000")
                    except (ValueError, InvalidOperation):
                        total_fees = Decimal("5000")

                    # Parse first installment
                    try:
                        first_installment = (
                            Decimal(str(first_installment_val)) if first_installment_val else Decimal("0")
                        )
                    except (ValueError, InvalidOperation):
                        first_installment = Decimal("0")

                    # Validate mobile
                    mobile_str = str(mobile_own).strip()
                    if not is_valid_mobile(mobile_str):
                        error_detail = "Mobile must be 10 digits starting with 6, 7, 8, or 9"
                        error_rows.append(
                            {"row": row_num, "field": "Mobile (Own)", "value": mobile_own, "error": error_detail}
                        )
                        continue

                    # Validate pin code
                    pin_str = str(pin_code).strip()
                    if not is_valid_pincode(pin_str):
                        error_detail = "Pin code must be exactly 6 digits"
                        error_rows.append(
                            {"row": row_num, "field": "Pin Code", "value": pin_code, "error": error_detail}
                        )
                        continue

                    # Check if student already exists (by full_name and mobile)
                    existing = AdmittedStudent.objects.filter(
                        full_name__iexact=full_name, mobile_own=mobile_str
                    ).exists()

                    if existing:
                        error_rows.append(
                            {
                                "row": row_num,
                                "field": "Full Name + Mobile",
                                "value": f"{full_name} ({mobile_own})",
                                "error": "Duplicate student - already exists in database",
                            }
                        )
                        continue

                    # Create admission record
                    admission = AdmittedStudent.objects.create(
                        student_name=student_name[:100],
                        father_name=father_name[:100],
                        surname=surname[:100],
                        mother_name=mother_name[:100] if mother_name else "",
                        full_name=full_name[:300],
                        date_of_birth=dob,
                        mobile_own=mobile_str,
                        parent_mobile=str(parent_mobile).strip() if parent_mobile else "",
                        gender=gender,
                        marital_status=marital_status,
                        course=course,
                        custom_course=custom_course[:100] if custom_course else "",
                        educational_qualification=educational_qualification[:200],
                        address=address,
                        city=city[:100],
                        tehsil_block=tehsil_block[:100],
                        district=district[:100],
                        pin_code=pin_str,
                        batch_month=str(batch_month).strip() if batch_month else None,
                        batch_year=str(batch_year).strip() if batch_year else None,
                        total_fees=total_fees,
                        paid_fees=Decimal("0"),
                        admission_date=admission_date,
                    )

                    # Create FeePayment receipt if first_installment > 0 and payment_mode is valid
                    valid_payment_modes = ["Cash", "UPI", "Card", "Bank Transfer"]
                    if first_installment > Decimal("0") and payment_mode in valid_payment_modes:
                        # Generate unique receipt_no: REC-DDMMYYYY-XXXXX
                        receipt_prefix = admission_date.strftime("%d%m%Y")
                        receipt_suffix = str(uuid.uuid4().hex[:5]).upper()
                        receipt_no = f"REC-{receipt_prefix}-{receipt_suffix}"

                        # Create FeePayment record
                        FeePayment.objects.create(
                            receipt_no=receipt_no,
                            student=admission,
                            amount=first_installment,
                            payment_mode=payment_mode,
                            payment_date=admission_date,
                            total_fees_at_payment=total_fees,
                            paid_before_this=Decimal("0"),
                            remaining_after_this=total_fees - first_installment,
                        )

                        # Update admission paid_fees to reflect this payment
                        admission.paid_fees = first_installment
                        admission.save(update_fields=["paid_fees"])

                    imported_count += 1

                except Exception as e:
                    error_rows.append({"row": row_num, "error": str(e)[:100]})
                    continue

        # Prepare response
        success_message = f"✅ Successfully imported {imported_count} admission(s)"

        if error_rows:
            error_message = f"⚠️ {len(error_rows)} row(s) had errors"
            return JsonResponse(
                {
                    "success": True,
                    "imported_count": imported_count,
                    "error_count": len(error_rows),
                    "message": success_message,
                    "warning": error_message,
                    "errors": error_rows[:10],  # Return first 10 errors
                }
            )
        else:
            return JsonResponse({"success": True, "imported_count": imported_count, "message": success_message})

    except Exception as e:
        return JsonResponse({"success": False, "error": f"Server error: {str(e)}"}, status=500)


# ================= IMPORT STUDENT PHOTOS FROM ZIP =================
@login_required
@staff_member_required
@csrf_protect
@require_http_methods(["POST"])
def import_student_photos_zip(request):
    """Import student photos from ZIP file - matches by surname and name"""
    try:
        zip_file = request.FILES.get("zip_file")

        if not zip_file:
            return JsonResponse({"success": False, "message": "Please select a ZIP file"}, status=400)

        matched = 0
        mismatched = []
        errors = []

        with zipfile.ZipFile(zip_file, "r") as zip_ref:
            for file_info in zip_ref.filelist:
                # Skip directories and hidden files
                if file_info.is_dir() or file_info.filename.startswith("__"):
                    continue

                filename = os.path.basename(file_info.filename)
                name_without_ext = os.path.splitext(filename)[0].strip()

                # Parse surname and name from filename
                # Expected format: "Surname Name.jpg" or "surname_name.jpg"
                # Try space separator first, then underscore
                parts = None
                if " " in name_without_ext:
                    parts = name_without_ext.split(" ", 1)
                elif "_" in name_without_ext:
                    parts = name_without_ext.split("_", 1)
                else:
                    # Single name, try to match with student_name
                    students = AdmittedStudent.objects.filter(student_name__icontains=name_without_ext)
                    if not students.exists():
                        mismatched.append(f"{filename} (Expected: 'Surname Name.jpg')")
                        continue
                    if students.count() > 1:
                        errors.append(f"{filename}: Multiple students match. Use 'Surname Name' format")
                        continue
                    parts = None

                if parts and len(parts) == 2:
                    surname, student_name = parts[0].strip(), parts[1].strip()

                    # Search for student by surname AND name (case-insensitive)
                    students = AdmittedStudent.objects.filter(
                        surname__icontains=surname, student_name__icontains=student_name
                    )
                elif parts is None and len(mismatched) > 0:
                    # Already handled above, skip
                    continue
                else:
                    mismatched.append(f"{filename} (Expected: 'Surname Name.jpg')")
                    continue

                if not students.exists():
                    mismatched.append(f"{filename} (No match: {surname} {student_name})")
                    continue

                if students.count() > 1:
                    errors.append(f"{filename}: {students.count()} students match. Be more specific")
                    continue

                student = students.first()

                # Read and validate image
                try:
                    file_content = zip_ref.read(file_info.filename)

                    # Validate it's a real image
                    img = Image.open(BytesIO(file_content))
                    img.verify()

                    # Save photo
                    file_extension = os.path.splitext(filename)[1].lower()
                    photo_name = f"student_photos/{student.id}_{student.surname}_{student.student_name}{file_extension}"

                    # Remove old photo if exists
                    if student.photo:
                        student.photo.delete()

                    # Save new photo
                    from django.core.files.base import ContentFile

                    student.photo.save(photo_name, ContentFile(file_content), save=True)
                    matched += 1

                except Exception as e:
                    errors.append(f"{filename}: Invalid image file - {str(e)[:50]}")

        return JsonResponse(
            {
                "success": True,
                "message": f"{matched} photos imported successfully",
                "matched": matched,
                "mismatched": mismatched,
                "errors": errors,
                "mismatched_count": len(mismatched),
                "error_count": len(errors),
            }
        )

    except zipfile.BadZipFile:
        return JsonResponse(
            {
                "success": False,
                "message": "Invalid ZIP file. Please upload a valid ZIP file.",
                "matched": 0,
                "mismatched": [],
                "errors": [],
            },
            status=400,
        )
    except Exception as e:
        return JsonResponse(
            {"success": False, "message": f"Error: {str(e)}", "matched": 0, "mismatched": [], "errors": []}, status=500
        )


# ================= ADD COURSE TO DATABASE (SINGLE DEFINITION) =================
@login_required
@require_http_methods(["POST"])
@csrf_protect
def add_course_ajax(request):
    """Add a new course via AJAX"""
    try:
        data = json.loads(request.body)
        course_name = data.get("course_name", "").strip()

        # Validate using CourseForm
        form = CourseForm(data={"name": course_name, "duration": "To be defined"})

        if form.is_valid():
            course = form.save()

            # ✅ IMPORTANT: Clear the course cache so new course appears everywhere immediately
            from django.core.cache import cache

            cache.delete("courses_list")

            return JsonResponse(
                {
                    "success": True,
                    "message": f'Course "{course_name}" added successfully!',
                    "course_id": course.id,
                    "course_name": course.name,
                },
                status=201,
            )
        else:
            errors = [str(error) for field_errors in form.errors.values() for error in field_errors]
            return JsonResponse(
                {"success": False, "message": " | ".join(errors) if errors else "Invalid course data"}, status=400
            )

    except json.JSONDecodeError:
        return JsonResponse({"success": False, "message": "Invalid JSON data"}, status=400)
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Error: {str(e)}"}, status=500)


# ================= ADMITTED STUDENTS LIST =================
@login_required
def admitted_students(request):
    """Render the admitted student list with filters, metrics, and pagination."""
    search = request.GET.get("search", "")
    month = request.GET.get("month", "")
    year = request.GET.get("year", "")
    course = request.GET.get("course", "")
    batch_month = request.GET.get("batch_month", "")  # NEW
    batch_year = request.GET.get("batch_year", "")  # NEW
    view_mode = "compact" if request.GET.get("view") == "compact" else "default"

    # Optimized query with only required fields to reduce database hits
    students = AdmittedStudent.objects.only(
        "id",
        "student_id",
        "full_name",
        "student_name",
        "father_name",
        "surname",
        "mobile_own",
        "parent_mobile",
        "course",
        "custom_course",
        "admission_date",
        "city",
        "total_fees",
        "paid_fees",
        "photo",
        "batch_month",
        "batch_year",
        "batch_status",
        "theory_batch_time",
        "practical_batch_time",
    )

    if search:
        students = students.filter(
            Q(full_name__icontains=search) | Q(student_name__icontains=search) | Q(mobile_own__icontains=search)
        )

    if month:
        students = students.filter(admission_date__month=month)

    if year:
        students = students.filter(admission_date__year=year)

    if course:
        students = students.filter(course=course)

    # NEW: Batch filters
    if batch_month:
        students = students.filter(batch_month=batch_month)

    if batch_year:
        students = students.filter(batch_year=batch_year)

    # Handle sorting
    sort = request.GET.get("sort", "")
    if sort == "name_asc":
        # Sort A-Z: Surname → Student Name → Father Name
        students = students.order_by("surname", "student_name", "father_name")
    elif sort == "name_desc":
        # Sort Z-A: Surname → Student Name → Father Name (descending)
        students = students.order_by("-surname", "-student_name", "-father_name")
    elif sort == "course":
        students = students.order_by("course")
    elif sort == "batch":
        students = students.order_by("batch_year", "batch_month")
    elif sort == "remaining_asc":
        # Sort by remaining fees (total_fees - paid_fees) in ascending order
        students = students.annotate(remaining=F("total_fees") - F("paid_fees")).order_by("remaining")
    elif sort == "remaining_desc":
        # Sort by remaining fees (total_fees - paid_fees) in descending order
        students = students.annotate(remaining=F("total_fees") - F("paid_fees")).order_by("-remaining")
    else:
        # Default sorting by surname and student name
        students = students.order_by("surname", "student_name")

    # Optimized: Use values_list instead of full objects for dropdown data
    available_years = (
        AdmittedStudent.objects.annotate(year=ExtractYear("admission_date"))
        .values_list("year", flat=True)
        .distinct()
        .order_by("-year")
    )

    # NEW: Get available batch months and years (optimized)
    available_batch_months = (
        AdmittedStudent.objects.exclude(batch_month__isnull=True)
        .exclude(batch_month="")
        .values_list("batch_month", flat=True)
        .distinct()
        .order_by("batch_month")
    )

    available_batch_years = (
        AdmittedStudent.objects.exclude(batch_year__isnull=True)
        .exclude(batch_year="")
        .values_list("batch_year", flat=True)
        .distinct()
        .order_by("-batch_year")
    )

    # Optimized: Cache courses (small table, rarely changes)
    all_courses = get_cached_courses()

    # Compute dashboard metrics against the full filtered result before paginating the visible table.
    students = list(students)
    today = timezone.localdate()
    total_students = len(students)
    total_fee_value = Decimal("0")
    total_paid_value = Decimal("0")
    total_remaining_value = Decimal("0")
    active_students_count = 0
    completed_students_count = 0
    pending_fee_count = 0
    dropout_count = 0
    batch_counter = Counter()
    course_counter = Counter()
    month_counter = Counter()

    for student in students:
        course_display = (
            student.custom_course if student.course == "Other" and student.custom_course else student.course
        ) or "Not Assigned"
        remaining_fees = student.remaining_fees or Decimal("0")
        paid_fees = student.paid_fees or Decimal("0")
        total_fees = student.total_fees or Decimal("0")
        batch_label = (
            f"{student.batch_month or 'Batch'} {student.batch_year or ''}".strip()
            if (student.batch_month or student.batch_year)
            else "Unassigned"
        )
        if total_fees > 0:
            payment_progress = max(0, min(100, round((paid_fees / total_fees) * Decimal("100"))))
        else:
            payment_progress = 0

        if remaining_fees <= 0:
            fee_state = "paid"
        elif paid_fees > 0:
            fee_state = "partial"
        else:
            fee_state = "unpaid"

        if student.batch_status != "active":
            lifecycle_stage = "Completed"
        elif (today - student.admission_date).days <= 15:
            lifecycle_stage = "Newly Admitted"
        elif remaining_fees > 0:
            lifecycle_stage = "Fee Pending"
        else:
            lifecycle_stage = "Active"

        student.course_display = course_display
        student.batch_label = batch_label
        student.payment_progress = payment_progress
        student.fee_state = fee_state
        student.lifecycle_stage = lifecycle_stage
        student.lifecycle_slug = lifecycle_stage.lower().replace(" ", "-")
        student.is_overdue_fee = remaining_fees > 0 and (today - student.admission_date).days > 30
        student.mkcl_fee_paid = "Yes" if paid_fees > 0 else "Pending"
        student.profit_estimate = max(Decimal("0"), paid_fees * Decimal("0.22"))

        total_fee_value += total_fees
        total_paid_value += paid_fees
        total_remaining_value += remaining_fees
        if student.batch_status == "active":
            active_students_count += 1
        else:
            completed_students_count += 1
        if remaining_fees > 0:
            pending_fee_count += 1
        if student.batch_status not in {"active", "ended"}:
            dropout_count += 1
        if batch_label != "Unassigned":
            batch_counter[batch_label] += 1
        course_counter[course_display] += 1
        month_counter[student.admission_date.strftime("%b")] += 1

    fee_recovery_rate = round((total_paid_value / total_fee_value) * 100, 1) if total_fee_value > 0 else 0
    retention_rate = round((active_students_count / total_students) * 100, 1) if total_students else 0
    average_ticket = round(float(total_fee_value / total_students), 2) if total_students else 0
    batch_occupancy_rate = round((sum(batch_counter.values()) / total_students) * 100, 1) if total_students else 0
    top_courses = course_counter.most_common(5)
    top_batches = batch_counter.most_common(5)
    admission_trends = [{"label": label, "count": count} for label, count in month_counter.items()]
    admission_trends.sort(key=lambda item: dt.strptime(item["label"], "%b").month)

    paginator = Paginator(students, max(total_students, 1))
    page_obj = paginator.get_page(1)

    return render(
        request,
        "core/admitted_students.html",
        {
            "students": page_obj.object_list,
            "page_obj": page_obj,
            "search": search,
            "month": month,
            "year": year,
            "course": course,
            "batch_month": batch_month,  # NEW
            "batch_year": batch_year,  # NEW
            "view_mode": view_mode,
            "sort": sort,  # NEW - Sorting parameter
            "available_years": available_years,
            "available_batch_months": available_batch_months,  # NEW
            "available_batch_years": available_batch_years,  # NEW
            "active_page": "admitted_students",
            "all_courses": all_courses,
            "time_slots": get_cached_time_slots(),
            "time_slot_display_map": TIME_SLOT_DISPLAY_MAP,
            "dashboard_metrics": {
                "total_students": total_students,
                "active_students": active_students_count,
                "pending_fees": pending_fee_count,
                "revenue_collected": total_paid_value,
                "revenue_pending": total_remaining_value,
                "completed_students": completed_students_count,
                "dropout_count": dropout_count,
                "batch_occupancy": batch_occupancy_rate,
                "fee_recovery_rate": fee_recovery_rate,
                "retention_rate": retention_rate,
                "average_ticket": average_ticket,
                "expected_revenue": total_fee_value,
            },
            "top_courses": top_courses,
            "top_batches": top_batches,
            "admission_trends": admission_trends,
        },
    )


# ================= STUDENT DETAIL (ADMITTED) =================
@login_required
def student_detail_admitted(request, student_id):
    """Get admitted student details via AJAX"""
    try:
        student = AdmittedStudent.objects.prefetch_related("fee_payments").get(id=student_id)

        # Get payment history
        fee_payments = student.fee_payments.order_by("payment_date")
        payment_history = []

        for payment in fee_payments:
            # Note: payment_date is now a DateField (no time component)
            payment_history.append(
                {
                    "id": payment.id,
                    "payment_date": payment.payment_date.strftime("%d-%m-%Y") if payment.payment_date else "",
                    "payment_time": "",  # No time for DateField
                    "amount": float(payment.amount),
                    "payment_mode": payment.payment_mode,
                    "receipt_no": payment.receipt_no or "",
                    "remaining_after": float(payment.remaining_after_this),
                }
            )

        # ✅ Include ALL fields from the form
        data = {
            "id": student.id,
            "student_name": student.student_name,
            "father_name": student.father_name,
            "surname": student.surname,
            "mother_name": student.mother_name,
            "full_name": student.full_name,
            "date_of_birth": student.date_of_birth.strftime("%Y-%m-%d") if student.date_of_birth else "",
            "gender": student.gender,
            "marital_status": student.marital_status,
            "mobile_own": student.mobile_own,
            "parent_mobile": student.parent_mobile or "",
            "course": student.course,
            "custom_course": student.custom_course or "",
            "educational_qualification": student.educational_qualification,
            "batch_month": student.batch_month or "",
            "batch_year": student.batch_year or "",
            "batch_display": student.batch_display or "Not Assigned",
            "batch_status": student.batch_status,
            "batch_end_date": student.batch_end_date.strftime("%Y-%m-%d") if student.batch_end_date else "",
            "batch_restored_date": (
                student.batch_restored_date.strftime("%Y-%m-%d") if student.batch_restored_date else ""
            ),
            "theory_batch_time": student.theory_batch_time or "",
            "practical_batch_time": student.practical_batch_time or "",
            "address": student.address,
            "city": student.city,
            "tehsil_block": student.tehsil_block,
            "district": student.district,
            "pin_code": student.pin_code,
            "photo": student.photo.url if student.photo else None,
            "total_fees": float(student.total_fees),
            "paid_fees": float(student.paid_fees),
            "remaining_fees": float(student.remaining_fees),
            "admission_date": student.admission_date.strftime("%Y-%m-%d") if student.admission_date else "",
            "payment_history": payment_history,
        }

        return JsonResponse(data)
    except AdmittedStudent.DoesNotExist:
        return JsonResponse({"error": "Student not found"}, status=404)
    except Exception as e:
        print(f"Error: {str(e)}")
        return JsonResponse({"error": str(e)}, status=500)


# ================= UPDATE STUDENT (ADMITTED) =================
@login_required
def update_student_admitted(request, student_id):
    """Apply AJAX updates to an admitted student record."""
    if request.method == "POST":
        student = get_object_or_404(AdmittedStudent, id=student_id)
        original_total_fees = student.total_fees

        # Handle both JSON and POST form data
        if request.content_type == "application/json":
            try:
                data = json.loads(request.body)
            except json.JSONDecodeError:
                return JsonResponse({"success": False, "error": "Invalid JSON"}, status=400)
        else:
            data = request.POST.dict()

        # Only update fields that are provided
        if "student_name" in data:
            student.student_name = data.get("student_name")
        if "father_name" in data:
            student.father_name = data.get("father_name")
        if "surname" in data:
            student.surname = data.get("surname")
        if "mother_name" in data:
            student.mother_name = data.get("mother_name")
        if "full_name" in data:
            student.full_name = data.get("full_name")
        if "date_of_birth" in data:
            student.date_of_birth = data.get("date_of_birth")
        if "admission_date" in data:
            student.admission_date = data.get("admission_date")
        if "mobile_own" in data:
            student.mobile_own = data.get("mobile_own")
        if "parent_mobile" in data:
            student.parent_mobile = data.get("parent_mobile")
        if "gender" in data:
            student.gender = data.get("gender")
        if "marital_status" in data:
            student.marital_status = data.get("marital_status")
        if "course" in data:
            student.course = data.get("course")
        if "custom_course" in data:
            student.custom_course = data.get("custom_course")
        if "educational_qualification" in data:
            student.educational_qualification = data.get("educational_qualification")
        if "address" in data:
            student.address = data.get("address")
        if "city" in data:
            student.city = data.get("city")
        if "tehsil_block" in data:
            student.tehsil_block = data.get("tehsil_block")
        if "district" in data:
            student.district = data.get("district")
        if "pin_code" in data:
            student.pin_code = data.get("pin_code")

        # Update batch information
        if "batch_month" in data:
            student.batch_month = data.get("batch_month", "")
        if "batch_year" in data:
            student.batch_year = data.get("batch_year", "")
        if "theory_batch_time" in data:
            theory_batch_time = (data.get("theory_batch_time") or "").strip()
            try:
                _validate_single_batch_capacity(
                    student=student,
                    field_name="theory_batch_time",
                    new_value=theory_batch_time,
                )
            except ValueError as exc:
                return JsonResponse({"success": False, "error": str(exc)}, status=400)
            student.theory_batch_time = theory_batch_time or None
        if "practical_batch_time" in data:
            practical_batch_time = (data.get("practical_batch_time") or "").strip()
            try:
                _validate_single_batch_capacity(
                    student=student,
                    field_name="practical_batch_time",
                    new_value=practical_batch_time,
                )
            except ValueError as exc:
                return JsonResponse({"success": False, "error": str(exc)}, status=400)
            student.practical_batch_time = practical_batch_time or None

        total_fees_changed = False

        # Update fees information
        if "total_fees" in data:
            total_fees_raw = str(data.get("total_fees", "")).strip()
            if total_fees_raw != "":
                try:
                    new_total_fees = Decimal(total_fees_raw)
                except (InvalidOperation, TypeError, ValueError):
                    return JsonResponse({"success": False, "error": "Invalid total fees amount"}, status=400)

                if new_total_fees < 0:
                    return JsonResponse({"success": False, "error": "Total fees cannot be negative"}, status=400)

                if new_total_fees < (student.paid_fees or Decimal("0")):
                    return JsonResponse(
                        {"success": False, "error": "Total fees cannot be less than paid fees"}, status=400
                    )

                student.total_fees = new_total_fees
                total_fees_changed = new_total_fees != original_total_fees

        # Handle photo removal flag
        if data.get("remove_photo") == "true":
            print("[PHOTO DEBUG] Remove photo flag detected")
            if student.photo:
                try:
                    print(f"[PHOTO DEBUG] Deleting photo: {student.photo}")
                    student.photo.delete()
                    student.photo = None
                except Exception as e:
                    print(f"[PHOTO DEBUG] Error deleting photo: {e}")
        # Handle photo upload (only if not removing)
        elif "photo" in request.FILES:
            photo_file = request.FILES["photo"]
            print(f"[PHOTO DEBUG] Photo file found: {photo_file.name} ({photo_file.size} bytes)")
            # Delete old photo if it exists
            if student.photo:
                try:
                    print(f"[PHOTO DEBUG] Deleting old photo: {student.photo}")
                    student.photo.delete()
                except Exception as e:
                    print(f"[PHOTO DEBUG] Error deleting old photo: {e}")
            # Save new photo
            student.photo = photo_file
            print("[PHOTO DEBUG] Photo assigned to student")
        else:
            print(
                "[PHOTO DEBUG] No photo action. "
                f'Remove flag: {data.get("remove_photo")}, '
                f"Files: {list(request.FILES.keys())}"
            )

        with transaction.atomic():
            student.save()

            # Keep historical payment snapshots aligned with the updated course fee.
            if total_fees_changed:
                running_paid_total = Decimal("0")
                payments = FeePayment.objects.filter(student=student).order_by("payment_date", "created_at", "id")
                for payment in payments:
                    payment.total_fees_at_payment = student.total_fees
                    payment.paid_before_this = running_paid_total
                    running_paid_total += payment.amount
                    payment.remaining_after_this = student.total_fees - running_paid_total
                    payment.save(update_fields=["total_fees_at_payment", "paid_before_this", "remaining_after_this"])

            # Update StudentFinanceDetail if total_fees changed
            if total_fees_changed:
                finance_detail, created = StudentFinanceDetail.objects.get_or_create(student=student)
                finance_detail.save()

        log_audit_event(
            action="student.updated",
            actor=request.user,
            target=student,
            request=request,
            metadata={
                "updated_fields": sorted(data.keys()),
                "total_fees_changed": total_fees_changed,
            },
        )

        return JsonResponse(
            {
                "success": True,
                "total_fees": float(student.total_fees),
                "paid_fees": float(student.paid_fees or 0),
                "remaining_fees": float(student.remaining_fees),
            }
        )

    return JsonResponse({"success": False, "error": "Invalid request"})


# ================= SEARCH ADMITTED STUDENTS (AJAX) =================
@login_required
def search_admitted_students(request):
    """Return admitted students for the global student search autocomplete."""
    query = request.GET.get("q", "").strip()

    if len(query) < 3:
        return JsonResponse({"students": []})

    students = AdmittedStudent.objects.filter(
        Q(full_name__icontains=query) | Q(student_name__icontains=query) | Q(mobile_own__icontains=query)
    ).order_by("-admission_date")[:10]

    students_data = []
    for student in students:
        course_name = student.custom_course if student.course == "Other" and student.custom_course else student.course
        students_data.append(
            {
                "id": student.id,
                "full_name": escape(
                    student.full_name or f"{student.student_name} {student.father_name} {student.surname}"
                ),
                "mobile_own": escape(student.mobile_own or ""),
                "course": escape(course_name or ""),
            }
        )

    return JsonResponse({"students": students_data})


# ================= FEES PAYMENT PAGE =================
@login_required
def fees_payment(request):
    """Render the fee payment entry page."""
    return render(request, "core/fees_payment.html", {"active_page": "fees_payment"})


# ================= SEARCH STUDENTS FOR FEES PAYMENT =================
@login_required
def search_students_for_payment(request):
    """Return matching students for the fee payment typeahead search."""
    query = request.GET.get("q", "").strip()

    if len(query) < 2:
        return JsonResponse({"students": []})

    students = (
        AdmittedStudent.objects.filter(
            Q(full_name__icontains=query) | Q(student_name__icontains=query) | Q(mobile_own__icontains=query)
        )
        .only("id", "full_name", "mobile_own", "course", "custom_course")
        .order_by("full_name")[:10]
    )

    students_data = []
    for student in students:
        course_name = student.custom_course if student.course == "Other" and student.custom_course else student.course
        students_data.append(
            {
                "id": student.id,
                "full_name": escape(student.full_name or ""),
                "mobile_own": escape(student.mobile_own or ""),
                "course": escape(course_name or ""),
            }
        )

    return JsonResponse({"students": students_data})


# ================= SUBMIT FEE PAYMENT - FIXED VERSION WITH BATCH =================
@login_required
@csrf_protect
@require_http_methods(["POST"])
def submit_fee_payment(request):
    """Validate and persist a learner fee payment, then return receipt data."""
    if request.method == "POST":
        try:
            # Get form data
            student_id = request.POST.get("student_id")
            amount = request.POST.get("amount")
            payment_mode = request.POST.get("payment_mode")
            payment_date = request.POST.get("payment_date")
            remarks = request.POST.get("remarks", "")

            logger.debug(
                "Received payment data for student_id=%s, amount=%s, payment_mode=%s, payment_date=%s",
                student_id,
                amount,
                payment_mode,
                payment_date,
            )

            # Validate with FeePaymentForm
            form_data = {
                "student": student_id,
                "amount": amount,
                "payment_mode": payment_mode,
                "payment_date": payment_date,
                "remarks": remarks,
            }
            form = FeePaymentForm(form_data)

            if not form.is_valid():
                errors = [str(error) for field_errors in form.errors.values() for error in field_errors]
                return JsonResponse(
                    {"success": False, "error": " | ".join(errors) if errors else "Invalid payment data"}, status=400
                )

            # Parse payment_date string (YYYY-MM-DD format) to date object
            try:
                from datetime import datetime

                payment_date_obj = datetime.strptime(payment_date, "%Y-%m-%d").date()
                payment_date_formatted = payment_date_obj.strftime("%d-%m-%Y")
            except ValueError:
                return JsonResponse({"success": False, "error": "Invalid payment date format"}, status=400)

            # Convert amount to Decimal
            try:
                amount = Decimal(str(amount).strip())
                # Validate amount is positive
                if amount <= 0:
                    raise ValueError("Amount must be greater than zero")
                # Validate amount doesn't exceed maximum (₹10 million)
                if amount > Decimal("10000000"):
                    raise ValueError("Amount exceeds maximum limit (₹10,000,000)")
                # Quantize to 2 decimal places (paise)
                amount = amount.quantize(Decimal("0.01"))
            except (ValueError, TypeError) as e:
                return JsonResponse({"success": False, "error": f"Invalid amount: {str(e)}"}, status=400)

            # Use atomic transaction
            with transaction.atomic():
                # Get student with lock
                try:
                    student = AdmittedStudent.objects.select_for_update().get(id=student_id)
                except AdmittedStudent.DoesNotExist:
                    return JsonResponse({"success": False, "error": "Student not found"}, status=404)

                # Check if amount exceeds remaining fees
                if amount > student.remaining_fees:
                    remaining_fees = student.remaining_fees
                    return JsonResponse(
                        {
                            "success": False,
                            "error": (
                                f"Payment amount (₹{amount}) cannot exceed remaining fees " f"(₹{remaining_fees})"
                            ),
                        },
                        status=400,
                    )

                # Create payment record with user-selected payment date
                payment = FeePayment.objects.create(
                    student=student,
                    amount=amount,
                    payment_mode=payment_mode,
                    payment_date=payment_date_obj,
                    remarks=remarks,
                    total_fees_at_payment=student.total_fees,
                    paid_before_this=student.paid_fees,
                    remaining_after_this=student.total_fees - (student.paid_fees + amount),
                )

                # Update student's paid fees using F() for atomic increment
                student.paid_fees = F("paid_fees") + amount
                student.save(update_fields=["paid_fees"])

                # Prepare receipt data
                course_name = (
                    student.custom_course if student.course == "Other" and student.custom_course else student.course
                )

                # ✅ FIXED: Get batch information correctly
                batch_month = student.batch_month or ""
                batch_year = student.batch_year or ""

                # Create batch display string
                if batch_month and batch_year:
                    batch_display = f"{batch_month} {batch_year}"
                else:
                    batch_display = "Not Assigned"

                receipt_data = {
                    "receipt_no": payment.receipt_no,
                    "date": payment_date_formatted,
                    "time": "",  # No time for DateField
                    "student_name": student.formatted_full_name,
                    "course": course_name,
                    "batch": batch_display,  # ✅ CORRECTED: Use batch_display
                    "mobile": student.mobile_own,
                    "payment_mode": payment_mode,
                    "total_fees": f"{float(student.total_fees):.2f}",
                    "previous_paid": f"{float(payment.paid_before_this):.2f}",
                    "amount_paid": f"{float(amount):.2f}",
                    "remaining_fees": f"{float(payment.remaining_after_this):.2f}",
                    "amount_in_words": number_to_words(float(amount)),
                }

                logger.info("Payment recorded successfully for receipt %s", receipt_data["receipt_no"])

                response_dict = {
                    "success": True,
                    "message": "Payment recorded successfully",
                    "receipt": receipt_data,
                    "_debug": "This is from the fixed submit_fee_payment view"
                }
                return JsonResponse(response_dict)

        except Exception as e:
            logger.exception("Error in submit_fee_payment")
            return JsonResponse({"success": False, "error": f"Server error: {str(e)}"}, status=500)

    return JsonResponse({"success": False, "error": "Invalid request method. Use POST."}, status=405)


# ================= EXPORT STUDENTS TO EXCEL =================
@login_required
def export_students_excel(request):
    """Export admitted student rows as an Excel workbook."""
    students = AdmittedStudent.objects.all().order_by("-admission_date")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Admitted Students"

    headers = ["S.No", "Full Name", "Mobile", "Course", "Total Fees", "Paid Fees", "Remaining Fees", "Admission Date"]
    ws.append(headers)

    for i, s in enumerate(students, 1):
        course = s.custom_course if s.course == "Other" else s.course
        ws.append(
            [
                i,
                s.full_name,
                s.mobile_own,
                course,
                float(s.total_fees),
                float(s.paid_fees),
                float(s.remaining_fees),
                s.admission_date.strftime("%d-%m-%Y"),
            ]
        )

    file = BytesIO()
    wb.save(file)
    file.seek(0)

    response = HttpResponse(
        file.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename=students_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return response


# ================= RECEIPTS VIEW =================
@login_required
def receipts_view(request):
    """Main receipts page"""
    return render(request, "core/receipts.html", {"active_page": "receipts"})


# ================= UPDATE RECEIPT API =================
@login_required
def get_receipts(request):
    """API endpoint to get receipts with filters"""
    try:
        receipts = FeePayment.objects.select_related("student").all().order_by("-payment_date")

        # Prepare receipt data
        receipt_list = []
        for receipt in receipts:
            student = receipt.student

            # ✅ Get batch information
            batch_month = student.batch_month or ""
            batch_year = student.batch_year or ""

            if batch_month and batch_year:
                batch_display = f"{batch_month} {batch_year}"
            else:
                batch_display = "Not Assigned"

            # Get course name (custom if 'Other' selected)
            course_name = (
                student.custom_course if student.course == "Other" and student.custom_course else student.course
            )

            receipt_list.append(
                {
                    "id": receipt.id,
                    "receipt_no": receipt.receipt_no,
                    "student_name": student.student_name,
                    "surname": student.surname or "",
                    "father_name": student.father_name or "",
                    "course": course_name,
                    "batch": batch_display,  # ✅ ADDED BATCH
                    "batch_display": batch_display,  # ✅ Fallback
                    "mobile": student.mobile_own,
                    "payment_mode": receipt.payment_mode,
                    "payment_date": str(receipt.payment_date),  # Convert DateField to string (YYYY-MM-DD)
                    "payment_time": "",  # No time for DateField
                    "paid_fees": float(receipt.amount),
                    "paid_before_this": float(receipt.paid_before_this),
                    "total_fees": float(student.total_fees),
                    "remaining_fees": float(receipt.remaining_after_this),
                }
            )

        return JsonResponse({"success": True, "receipts": receipt_list})
    except Exception as e:
        logger.exception("Error in get_receipts")
        return JsonResponse({"success": False, "error": str(e)}, status=500)


# ================= UPDATE RECEIPT API =================
@login_required
@staff_member_required
@require_http_methods(["POST"])
@csrf_protect
def update_receipt(request, receipt_id):
    """API endpoint to update a receipt"""
    try:
        data = json.loads(request.body)

        payment = FeePayment.objects.get(id=receipt_id)

        # Update only allowed fields
        if "payment_date" in data:
            # Parse date string if it comes in YYYY-MM-DD format
            try:
                from datetime import datetime

                payment_date_str = data["payment_date"]
                if isinstance(payment_date_str, str):
                    payment.payment_date = datetime.strptime(payment_date_str, "%Y-%m-%d").date()
                else:
                    payment.payment_date = payment_date_str
            except (ValueError, TypeError):
                return JsonResponse({"success": False, "error": "Invalid payment date format"}, status=400)

        if "amount" in data or "paid_fees" in data:
            old_amount = payment.amount
            try:
                new_amount = Decimal(str(data.get("amount") or data.get("paid_fees", old_amount)).strip())
            except (ValueError, TypeError, InvalidOperation):
                return JsonResponse({"success": False, "error": "Invalid amount format"}, status=400)

            # Validate amount
            if new_amount <= 0:
                return JsonResponse({"success": False, "error": "Amount must be greater than zero"}, status=400)

            if new_amount > Decimal("10000000"):
                return JsonResponse({"success": False, "error": "Amount exceeds maximum limit"}, status=400)

            # Update student's paid fees with validation
            student = payment.student
            amount_difference = new_amount - old_amount
            new_paid_fees = student.paid_fees + amount_difference

            # Validate new amount doesn't exceed total fees
            if new_paid_fees < 0:
                return JsonResponse(
                    {"success": False, "error": "Cannot update: would result in negative paid fees"}, status=400
                )

            if new_paid_fees > student.total_fees:
                return JsonResponse({"success": False, "error": "Cannot update: would exceed total fees"}, status=400)

            with transaction.atomic():
                # Use F() for atomic update
                student.paid_fees = F("paid_fees") + amount_difference
                student.save(update_fields=["paid_fees"])

                payment.amount = new_amount
                payment.remaining_after_this = student.total_fees - new_paid_fees
                payment.save()

        return JsonResponse({"success": True, "message": "Receipt updated successfully"})

    except FeePayment.DoesNotExist:
        return JsonResponse({"success": False, "error": "Receipt not found"}, status=404)
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Invalid JSON data"}, status=400)
    except Exception as e:
        print(f"Error updating receipt: {str(e)}")
        return JsonResponse({"success": False, "error": f"Error: {str(e)}"}, status=500)


# ================= DELETE RECEIPT API =================
@login_required
@staff_member_required
@require_http_methods(["POST"])
def delete_receipt(request, receipt_id):
    """API endpoint to delete a receipt"""
    try:
        payment = FeePayment.objects.select_related("student").get(id=receipt_id)
        student = payment.student
        receipt_no = payment.receipt_no
        student_total_fees = student.total_fees or Decimal("0")

        with transaction.atomic():
            payment.delete()

            remaining_payments = list(
                FeePayment.objects.filter(student_id=student.id).order_by("payment_date", "created_at", "id")
            )

            running_paid = Decimal("0")
            payments_to_update = []

            for remaining_payment in remaining_payments:
                paid_before_this = running_paid
                running_paid += remaining_payment.amount or Decimal("0")
                remaining_after_this = max(Decimal("0"), student_total_fees - running_paid)

                if (
                    remaining_payment.paid_before_this != paid_before_this
                    or remaining_payment.remaining_after_this != remaining_after_this
                ):
                    remaining_payment.paid_before_this = paid_before_this
                    remaining_payment.remaining_after_this = remaining_after_this
                    payments_to_update.append(remaining_payment)

            if payments_to_update:
                FeePayment.objects.bulk_update(payments_to_update, ["paid_before_this", "remaining_after_this"])

            AdmittedStudent.objects.filter(id=student.id).update(paid_fees=running_paid)

        return JsonResponse({"success": True, "message": f"Receipt {receipt_no} deleted successfully"})

    except FeePayment.DoesNotExist:
        return JsonResponse({"success": False, "error": "Receipt not found"}, status=404)
    except Exception as e:
        logger.exception("Error deleting receipt %s", receipt_id)
        return JsonResponse({"success": False, "error": f"Error deleting receipt: {str(e)}"}, status=500)


# ================= EXPORT RECEIPTS API =================
@login_required
def export_receipts(request):
    """Export receipts to Excel"""
    try:
        search = request.GET.get("search", "")
        date_filter = request.GET.get("date", "")
        month = request.GET.get("month", "")
        year = request.GET.get("year", "")

        payments = FeePayment.objects.select_related("student").all().order_by("-payment_date")

        if search:
            payments = payments.filter(
                Q(student__full_name__icontains=search)
                | Q(student__mobile_own__icontains=search)
                | Q(receipt_no__icontains=search)
            )

        if date_filter:
            try:
                filter_date = datetime.strptime(date_filter, "%Y-%m-%d").date()
                payments = payments.filter(payment_date__date=filter_date)
            except ValueError:
                pass

        if month:
            try:
                payments = payments.filter(payment_date__month=int(month))
            except ValueError:
                pass

        if year:
            try:
                payments = payments.filter(payment_date__year=int(year))
            except ValueError:
                pass

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Payment Receipts"

        headers = [
            "Receipt No",
            "Student Name",
            "Mobile",
            "Course",
            "Payment Date",
            "Payment Mode",
            "Total Fees",
            "Paid Before",
            "Amount Paid",
            "Remaining Fees",
            "Remarks",
        ]

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_num, payment in enumerate(payments, 2):
            course_name = (
                payment.student.custom_course
                if payment.student.course == "Other" and payment.student.custom_course
                else payment.student.course
            )

            ws.cell(row=row_num, column=1).value = payment.receipt_no
            ws.cell(row=row_num, column=2).value = payment.student.full_name
            ws.cell(row=row_num, column=3).value = payment.student.mobile_own
            ws.cell(row=row_num, column=4).value = course_name
            ws.cell(row=row_num, column=5).value = payment.payment_date.strftime("%d-%m-%Y")
            ws.cell(row=row_num, column=6).value = payment.payment_mode
            ws.cell(row=row_num, column=7).value = float(payment.total_fees_at_payment)
            ws.cell(row=row_num, column=8).value = float(payment.paid_before_this)
            ws.cell(row=row_num, column=9).value = float(payment.amount)
            ws.cell(row=row_num, column=10).value = float(payment.remaining_after_this)
            ws.cell(row=row_num, column=11).value = payment.remarks or ""

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        response = HttpResponse(
            excel_file.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        filename = f'receipts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)


# ================= EXPORT ADMITTED STUDENTS TO EXCEL =================
@login_required
def export_admitted_students_excel(request):
    """Export filtered admitted students for back-office reporting."""
    search = request.GET.get("search", "")
    month = request.GET.get("month", "")
    year = request.GET.get("year", "")
    course = request.GET.get("course", "")

    students = AdmittedStudent.objects.all()

    if search:
        students = students.filter(
            Q(full_name__icontains=search) | Q(student_name__icontains=search) | Q(mobile_own__icontains=search)
        )

    if month:
        students = students.filter(admission_date__month=month)

    if year:
        students = students.filter(admission_date__year=year)

    if course:
        students = students.filter(course=course)

    students = students.order_by("-admission_date")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Admitted Students"

    headers = [
        "S.No",
        "Full Name",
        "Student Name",
        "Father Name",
        "Surname",
        "Mother Name",
        "Date of Birth",
        "Mobile (Own)",
        "Parent Mobile",
        "Gender",
        "Marital Status",
        "Course",
        "Batch Month",
        "Batch Year",
        "Educational Qualification",
        "Address",
        "City",
        "Tehsil/Block",
        "District",
        "Pin Code",
        "Total Fees (₹)",
        "Paid Fees First Installment",
        "Admission Date",
    ]

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_num, student in enumerate(students, 2):
        ws.cell(row=row_num, column=1).value = row_num - 1
        ws.cell(row=row_num, column=2).value = student.full_name
        ws.cell(row=row_num, column=3).value = student.student_name
        ws.cell(row=row_num, column=4).value = student.father_name
        ws.cell(row=row_num, column=5).value = student.surname
        ws.cell(row=row_num, column=6).value = student.mother_name
        ws.cell(row=row_num, column=7).value = (
            student.date_of_birth.strftime("%d-%m-%Y") if student.date_of_birth else ""
        )
        ws.cell(row=row_num, column=8).value = student.mobile_own
        ws.cell(row=row_num, column=9).value = student.parent_mobile or ""
        ws.cell(row=row_num, column=10).value = student.gender
        ws.cell(row=row_num, column=11).value = student.marital_status
        ws.cell(row=row_num, column=12).value = student.course
        ws.cell(row=row_num, column=13).value = student.batch_month or ""
        ws.cell(row=row_num, column=14).value = student.batch_year or ""
        ws.cell(row=row_num, column=15).value = student.educational_qualification
        ws.cell(row=row_num, column=16).value = student.address
        ws.cell(row=row_num, column=17).value = student.city
        ws.cell(row=row_num, column=18).value = student.tehsil_block
        ws.cell(row=row_num, column=19).value = student.district
        ws.cell(row=row_num, column=20).value = student.pin_code
        ws.cell(row=row_num, column=21).value = float(student.total_fees)
        ws.cell(row=row_num, column=22).value = float(student.paid_fees) if student.paid_fees else 0
        ws.cell(row=row_num, column=23).value = student.admission_date.strftime("%d-%m-%Y")

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    response = HttpResponse(
        excel_file.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    filename_parts = ["admitted_students"]
    if search:
        filename_parts.append(f"search_{search[:20]}")
    if course:
        filename_parts.append(f"{course}")
    if month:
        filename_parts.append(f"month_{month}")
    if year:
        filename_parts.append(f"{year}")
    filename_parts.append(datetime.now().strftime("%Y%m%d_%H%M%S"))

    filename = "_".join(filename_parts) + ".xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response


# ================= DELETE ADMITTED STUDENTS =================
@login_required
@staff_member_required
@require_http_methods(["POST"])
def delete_admitted_students(request):
    """Delete multiple admitted students at once"""
    try:
        data = json.loads(request.body)
        student_ids = data.get("student_ids", [])

        if not student_ids:
            return JsonResponse({"success": False, "error": "No students selected"})

        try:
            student_ids = [int(id) for id in student_ids]
        except (ValueError, TypeError):
            return JsonResponse({"success": False, "error": "Invalid student IDs"})

        students_to_delete = AdmittedStudent.objects.filter(id__in=student_ids)

        if not students_to_delete.exists():
            return JsonResponse({"success": False, "error": "No students found with the given IDs"})

        delete_count = students_to_delete.count()

        with transaction.atomic():
            # Remove uploaded photo files before the ORM cascades the student rows.
            students_with_photos = list(students_to_delete.exclude(photo="").exclude(photo__isnull=True))
            for student in students_with_photos:
                try:
                    student.photo.delete(save=False)
                except Exception:
                    logger.warning("Unable to delete photo for student_id=%s", student.id, exc_info=True)

            students_to_delete.delete()

        return JsonResponse(
            {
                "success": True,
                "deleted_count": delete_count,
                "message": f"Successfully deleted {delete_count} student(s)",
            }
        )

    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Invalid JSON data"}, status=400)
    except Exception as e:
        logger.exception("Bulk admitted student delete failed")
        return JsonResponse({"success": False, "error": f"An error occurred: {str(e)}"}, status=500)


# ================= DATABASE BACKUP =================
@login_required
@staff_member_required
def backup_page(request):
    """Display the backup and restore page"""
    context = {
        "active_page": "backup",
    }
    return render(request, "core/backup.html", context)


@login_required
@staff_member_required
@require_http_methods(["GET"])
def export_database(request):
    """Export database as SQLite file with photos"""
    try:
        import zipfile
        from io import BytesIO

        db_path = settings.DATABASES["default"]["NAME"]

        # Convert Path object to string
        db_path = str(db_path)

        # Check if file exists
        if not os.path.exists(db_path):
            return JsonResponse({"success": False, "error": "Database file not found"}, status=500)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"ssc_education_backup_{timestamp}.zip"

        # Create a ZIP file in memory
        zip_buffer = BytesIO()

        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            # Add database file to ZIP
            db_filename = os.path.basename(db_path)
            zip_file.write(db_path, arcname=f"database/{db_filename}")

            # Add student photos if they exist
            media_path = os.path.join(settings.BASE_DIR, "media", "student_photos")
            if os.path.exists(media_path):
                for root, dirs, files in os.walk(media_path):
                    for file in files:
                        file_path = os.path.join(root, file)
                        # Calculate relative path for ZIP
                        relative_path = os.path.relpath(file_path, settings.BASE_DIR)
                        zip_file.write(file_path, arcname=relative_path)

        # Prepare response
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer.getvalue(), content_type="application/zip")
        response["Content-Disposition"] = f'attachment; filename="{backup_name}"'

        return response

    except Exception as e:
        logger.exception("Export database error")
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@login_required
@staff_member_required
@require_http_methods(["POST"])
@csrf_protect
def import_database(request):
    """Import database and photos from backup ZIP file (merge/update instead of overwrite)"""
    if "database_file" not in request.FILES:
        return JsonResponse({"success": False, "error": "No file provided"}, status=400)

    uploaded_file = request.FILES["database_file"]

    # Validate file - now accepts ZIP or raw database files
    valid_extensions = ["db", "sqlite", "sqlite3", "zip"]
    file_extension = uploaded_file.name.split(".")[-1].lower()

    if file_extension not in valid_extensions:
        return JsonResponse(
            {"success": False, "error": "Invalid file type. Only .db, .sqlite, .sqlite3, or .zip files are allowed."},
            status=400,
        )

    max_size = 500 * 1024 * 1024  # 500 MB for ZIP files with photos
    if uploaded_file.size > max_size:
        return JsonResponse({"success": False, "error": "File too large. Maximum size is 500 MB."}, status=400)

    try:
        import sqlite3
        import tempfile
        import zipfile

        db_path = settings.DATABASES["default"]["NAME"]

        # Create backup of current database before importing
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"database_backup_before_import_{timestamp}.db"
        backup_path = os.path.join(settings.BASE_DIR, backup_name)
        shutil.copy2(db_path, backup_path)

        # Create backup of current photos before importing
        media_path = os.path.join(settings.BASE_DIR, "media", "student_photos")
        if os.path.exists(media_path):
            photos_backup_dir = os.path.join(settings.BASE_DIR, f"student_photos_backup_{timestamp}")
            shutil.copytree(media_path, photos_backup_dir)

        temp_db_path = None
        temp_dir = None

        try:
            # Check if uploaded file is a ZIP file
            if file_extension == "zip":
                # Extract ZIP file
                temp_dir = tempfile.mkdtemp()
                with zipfile.ZipFile(uploaded_file, "r") as zip_ref:
                    safe_extract_zip(zip_ref, temp_dir)

                # Find database file in extracted ZIP
                temp_db_path = None
                for root, dirs, files in os.walk(temp_dir):
                    for file in files:
                        if file.endswith((".db", ".sqlite", ".sqlite3")):
                            temp_db_path = os.path.join(root, file)
                            break
                    if temp_db_path:
                        break

                if not temp_db_path:
                    return JsonResponse({"success": False, "error": "No database file found in backup ZIP"}, status=400)

                # Extract and restore student photos if they exist in ZIP
                photos_count = 0
                for root, dirs, files in os.walk(temp_dir):
                    for file in files:
                        if "student_photos" in root:
                            src_file = os.path.join(root, file)

                            # Extract only the filename from the original path structure
                            # Get relative path after 'student_photos' folder
                            parts = src_file.split(os.sep)
                            try:
                                student_photos_idx = [i for i, p in enumerate(parts) if "student_photos" in p][-1]
                                # Get path relative to student_photos folder
                                relative_filename = os.sep.join(parts[student_photos_idx + 1:])
                                # Destination in the project's media folder
                                dst_file = os.path.join(settings.BASE_DIR, "media", "student_photos", relative_filename)
                            except (IndexError, ValueError):
                                # Fallback: just use the filename
                                dst_file = os.path.join(settings.BASE_DIR, "media", "student_photos", file)

                            # Create directory if needed
                            os.makedirs(os.path.dirname(dst_file), exist_ok=True)

                            # Copy file
                            shutil.copy2(src_file, dst_file)
                            photos_count += 1
                            print(f"Imported photo: {dst_file}")
            else:
                # Handle raw database file upload (no photos)
                photos_count = 0
                with tempfile.NamedTemporaryFile(delete=False, suffix=".db") as tmp_file:
                    tmp_file.write(uploaded_file.read())
                    temp_db_path = tmp_file.name

            # Merge data from imported database into current database
            current_conn = sqlite3.connect(db_path)
            imported_conn = sqlite3.connect(temp_db_path)

            imported_cursor = imported_conn.cursor()
            current_cursor = current_conn.cursor()

            # Disable foreign key constraints temporarily to allow merging in any order
            current_cursor.execute("PRAGMA foreign_keys = OFF")
            imported_cursor.execute("PRAGMA foreign_keys = OFF")

            # Get all table names from imported database
            imported_cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%';")
            all_tables = [
                (table_name,)
                for (table_name,) in imported_cursor.fetchall()
                if table_name in IMPORT_ALLOWED_TABLES and is_safe_sqlite_identifier(table_name)
            ]

            # Sort tables to ensure dependencies are met (parent tables first)
            # Priority order: AdmittedStudent first, then dependent tables
            priority_tables = IMPORT_PRIORITY_TABLES

            # Start with priority tables, then add remaining tables
            tables_ordered = []
            for table in priority_tables:
                if any(t[0] == table for t in all_tables):
                    tables_ordered.append((table,))

            # Add remaining tables not in priority list
            for (table_name,) in all_tables:
                if not any(t[0] == table_name for t in tables_ordered):
                    tables_ordered.append((table_name,))

            merged_count = 0
            skipped_count = 0
            processed_tables = []

            # Merge each table's data
            for (table_name,) in tables_ordered:
                try:
                    processed_tables.append(table_name)
                    # Get columns for the table
                    imported_cursor.execute(f"PRAGMA table_info({table_name})")
                    columns_info = imported_cursor.fetchall()
                    columns = [col[1] for col in columns_info]
                    if not columns or not all(is_safe_sqlite_identifier(col) for col in columns):
                        skipped_count += 1
                        continue
                    primary_keys = [col[1] for col in columns_info if col[5]]  # col[5] is pk flag

                    # Get all records from imported table
                    imported_cursor.execute(f"SELECT * FROM {table_name}")
                    rows = imported_cursor.fetchall()

                    # Special handling for FeePayment (receipts) table
                    if table_name == "core_feepayment":
                        print(f"Processing {len(rows)} receipt records from FeePayment table")
                        for row in rows:
                            row_dict = dict(zip(columns, row))
                            receipt_no = row_dict.get("receipt_no", "")

                            # Check if receipt with same receipt_no already exists
                            current_cursor.execute(f"SELECT id FROM {table_name} WHERE receipt_no = ?", (receipt_no,))
                            existing_receipt = current_cursor.fetchone()

                            if existing_receipt:
                                # Receipt number already exists - UPDATE the record with new data
                                receipt_id = existing_receipt[0]
                                set_clause = ", ".join([f"{col} = ?" for col in columns if col != "id"])
                                values = [row_dict[col] for col in columns if col != "id"]
                                values.append(receipt_id)

                                try:
                                    current_cursor.execute(f"UPDATE {table_name} SET {set_clause} WHERE id = ?", values)
                                    merged_count += 1
                                    print(f"Updated receipt: {receipt_no} (ID: {receipt_id})")
                                except sqlite3.IntegrityError as ie:
                                    print(f"Error updating receipt {receipt_no}: {str(ie)}")
                                    skipped_count += 1
                            else:
                                # Receipt number doesn't exist - INSERT as new record with same receipt number
                                placeholders = ", ".join(["?" for _ in columns])
                                try:
                                    current_cursor.execute(
                                        f"INSERT INTO {table_name} ({', '.join(columns)}) VALUES ({placeholders})",
                                        tuple(row),
                                    )
                                    merged_count += 1
                                    print(f"Inserted new receipt: {receipt_no}")
                                except sqlite3.IntegrityError as ie:
                                    print(f"Error inserting receipt {receipt_no}: {str(ie)}")
                                    skipped_count += 1
                    else:
                        # Standard processing for other tables
                        # Insert or update records in current database
                        for row in rows:
                            row_dict = dict(zip(columns, row))

                            # Check if record exists (for tables with primary keys)
                            if primary_keys:
                                pk_column = primary_keys[0]
                                current_cursor.execute(
                                    f"SELECT 1 FROM {table_name} WHERE {pk_column} = ?", (row_dict[pk_column],)
                                )
                                record_exists = current_cursor.fetchone() is not None

                                if record_exists:
                                    # Update existing record
                                    set_clause = ", ".join([f"{col} = ?" for col in columns if col != pk_column])
                                    values = [row_dict[col] for col in columns if col != pk_column]
                                    values.append(row_dict[pk_column])

                                    current_cursor.execute(
                                        f"UPDATE {table_name} SET {set_clause} WHERE {pk_column} = ?", values
                                    )
                                    merged_count += 1
                                else:
                                    # Insert new record
                                    placeholders = ", ".join(["?" for _ in columns])
                                    current_cursor.execute(
                                        f"INSERT INTO {table_name} ({', '.join(columns)}) VALUES ({placeholders})",
                                        tuple(row),
                                    )
                                    merged_count += 1
                            else:
                                # If no primary key, just insert
                                placeholders = ", ".join(["?" for _ in columns])
                                current_cursor.execute(
                                    f"INSERT INTO {table_name} ({', '.join(columns)}) VALUES ({placeholders})",
                                    tuple(row),
                                )
                                merged_count += 1

                except Exception as table_error:
                    print(f"Error merging table {table_name}: {str(table_error)}")
                    import traceback

                    traceback.print_exc()
                    skipped_count += 1
                    continue

            # Re-enable foreign key constraints before committing
            current_cursor.execute("PRAGMA foreign_keys = ON")

            # Get count of receipts in the database
            current_cursor.execute("SELECT COUNT(*) FROM core_feepayment")
            total_receipts = current_cursor.fetchone()[0]

            if not processed_tables:
                imported_conn.close()
                current_conn.close()
                return JsonResponse(
                    {
                        "success": False,
                        "error": (
                            "No supported data tables were found in the uploaded backup. "
                            "This backup may be from a different app version or database schema."
                        ),
                    },
                    status=400,
                )

            # Commit changes and close connections
            current_conn.commit()
            current_conn.close()
            imported_conn.close()

            # Prepare success message
            message = (
                f"Database updated successfully. "
                f"{merged_count} records merged/updated across "
                f"{len(processed_tables)} table(s), "
                f"{skipped_count} table(s)/row groups skipped. "
                f"Total receipts in database: {total_receipts}. "
                f"Backup saved as {backup_name}"
            )
            if file_extension == "zip" and photos_count > 0:
                message += f". {photos_count} student photos imported to media/student_photos folder."
            elif file_extension == "zip":
                message += ". No student photos found in backup."

            print(f"Import completed: {message}")

            return JsonResponse(
                {
                    "success": True,
                    "message": message,
                    "merged_count": merged_count,
                    "skipped_count": skipped_count,
                    "processed_tables": processed_tables,
                }
            )

        finally:
            # Clean up temporary files
            if temp_db_path and os.path.exists(temp_db_path):
                os.remove(temp_db_path)
            if temp_dir and os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)

    except Exception as e:
        logger.exception("Import database error")
        return JsonResponse({"success": False, "error": f"Error importing database: {str(e)}"}, status=500)


@login_required
def statistics_view(request):
    """Main statistics page with year selection"""
    selected_year = request.GET.get("year", "")

    # Get available years from AdmittedStudent
    available_years = (
        AdmittedStudent.objects.annotate(year=ExtractYear("admission_date"))
        .values_list("year", flat=True)
        .distinct()
        .order_by("-year")
    )

    # Get students for selected year
    students = AdmittedStudent.objects.all()
    if selected_year:
        students = students.filter(admission_date__year=selected_year)

    # Calculate total profit for selected year (or current year if no selection)
    total_profit = calculate_total_profit(students)

    # Calculate total profit for all years
    all_students = AdmittedStudent.objects.all()
    total_profit_all_years = calculate_total_profit(all_students)

    # Get current year
    current_year = datetime.now().year
    current_year_students = AdmittedStudent.objects.filter(admission_date__year=current_year)
    total_profit_current_year = calculate_total_profit(current_year_students)

    context = {
        "available_years": available_years,
        "selected_year": selected_year,
        "total_profit": total_profit,
        "total_profit_current_year": total_profit_current_year,
        "total_profit_all_years": total_profit_all_years,
        "current_year": current_year,
        "total_admitted": students.count(),
        "student_count": students.count(),
        "active_page": "statistics",
    }
    return render(request, "core/statistics.html", context)


@login_required
def student_finance_details(request):
    """Student Finance Details section with filtering and sorting"""
    selected_year = request.GET.get("year", "")
    search_query = request.GET.get("search", "")
    sort_by = request.GET.get("sort", "name")  # default sort by name
    course_filter = request.GET.get("course", "")
    batch_filter = request.GET.get("batch", "")

    # Get available years from AdmittedStudent
    available_years = (
        AdmittedStudent.objects.annotate(year=ExtractYear("admission_date"))
        .values_list("year", flat=True)
        .distinct()
        .order_by("-year")
    )

    # Get all available courses
    available_courses = AdmittedStudent.objects.values_list("course", flat=True).distinct().order_by("course")

    # Get all available batches
    available_batches = (
        AdmittedStudent.objects.values_list("batch_month", "batch_year")
        .distinct()
        .order_by("-batch_year", "-batch_month")
    )
    # Format batches as "Month Year"
    formatted_batches = []
    for month, year in available_batches:
        if month and year:
            formatted_batches.append(f"{month} {year}")

    # Get all admitted students with filters
    students = AdmittedStudent.objects.prefetch_related(
        Prefetch("fee_payments", queryset=FeePayment.objects.order_by("payment_date")),
        "finance_detail",
    )

    # Apply year filter
    if selected_year:
        students = students.filter(admission_date__year=selected_year)

    # Apply course filter
    if course_filter:
        students = students.filter(course=course_filter)

    # Apply batch filter
    if batch_filter:
        batch_parts = batch_filter.split()
        if len(batch_parts) >= 2:
            month = " ".join(batch_parts[:-1])  # All but last part
            year = batch_parts[-1]  # Last part
            students = students.filter(batch_month=month, batch_year=year)

    # Apply search filter (search by name or mobile)
    if search_query:
        students = students.filter(
            Q(full_name__icontains=search_query)
            | Q(student_name__icontains=search_query)
            | Q(mobile_own__icontains=search_query)
        )

    finance_data = []
    total_profit = Decimal("0.00")
    total_learner_paid = Decimal("0.00")
    total_mkcl_paid = Decimal("0.00")

    for student in students:
        # Keep finance rows consistent even when a student record predates finance detail creation.
        finance_detail = getattr(student, "finance_detail", None)
        if finance_detail is None:
            finance_detail, _created = StudentFinanceDetail.objects.get_or_create(
                student=student,
                defaults={
                    "first_installment": Decimal("0.00"),
                    "second_installment": Decimal("0.00"),
                    "third_installment": Decimal("0.00"),
                    "fourth_installment": Decimal("0.00"),
                    "fifth_installment": Decimal("0.00"),
                    "fees_paid_to_mkcl_1": Decimal("0.00"),
                    "fees_paid_to_mkcl_2": Decimal("0.00"),
                },
            )

        # Calculate totals from AdmittedStudent
        total_paid = student.paid_fees or Decimal("0.00")
        total_fees = student.total_fees or Decimal("0.00")
        balance_fees = total_fees - total_paid

        # Get course name for defaults logic
        course_name = student.custom_course if student.course == "Other" and student.custom_course else student.course

        # Calculate fees paid to MKCL - default to 0
        mkcl_1 = finance_detail.fees_paid_to_mkcl_1 or Decimal("0.00")
        mkcl_2 = finance_detail.fees_paid_to_mkcl_2 or Decimal("0.00")
        mkcl_3 = finance_detail.fees_paid_to_mkcl_3 or Decimal("0.00")

        mkcl_total = mkcl_1 + mkcl_2 + mkcl_3

        # Get fee payments for this student - ordered by payment_date (oldest first)
        fee_payments = list(student.fee_payments.all())

        # Extract installment amounts from FeePayment records (5 installments)
        first_inst = Decimal("0.00")
        second_inst = Decimal("0.00")
        third_inst = Decimal("0.00")
        fourth_inst = Decimal("0.00")
        fifth_inst = Decimal("0.00")

        if len(fee_payments) >= 1:
            first_inst = fee_payments[0].amount
        if len(fee_payments) >= 2:
            second_inst = fee_payments[1].amount
        if len(fee_payments) >= 3:
            third_inst = fee_payments[2].amount
        if len(fee_payments) >= 4:
            fourth_inst = fee_payments[3].amount
        if len(fee_payments) >= 5:
            fifth_inst = fee_payments[4].amount

        # Calculate profit as (Total Fees Paid By Learner) - (Total Fees Paid to MKCL)
        # Use the actual total_paid from AdmittedStudent, not individual installments
        profit = total_paid - mkcl_total
        total_profit += profit
        total_learner_paid += total_paid
        total_mkcl_paid += mkcl_total

        # Build payment history (ordered by payment_date, newest first for display)
        payment_history = []
        for payment in reversed(fee_payments):
            payment_history.append(
                {
                    "receipt_no": payment.receipt_no,
                    "amount": payment.amount,
                    "payment_date": payment.payment_date,
                    "payment_mode": payment.payment_mode,
                    "remarks": payment.remarks,
                    "paid_before": payment.paid_before_this,
                    "remaining_after": payment.remaining_after_this,
                }
            )

        finance_data.append(
            {
                "id": student.id,
                "sr_no": student.id,
                "learner_name": student.full_name or f"{student.surname} {student.student_name} {student.father_name}",
                "student_id": student.id,  # Using student ID as identifier
                "mobile_no": student.mobile_own,
                "batch": student.batch_display,
                "course": course_name,
                "first_inst": first_inst,
                "second_inst": second_inst,
                "third_inst": third_inst,
                "fourth_inst": fourth_inst,
                "fifth_inst": fifth_inst,
                "total_paid": total_paid,
                "total_fees": total_fees,
                "balance_fees": balance_fees,
                "mkcl_1": mkcl_1,
                "mkcl_2": mkcl_2,
                "mkcl_3": mkcl_3,
                "mkcl_total": mkcl_total,
                "profit": profit,
                "payment_history": payment_history,
            }
        )

    # Apply sorting
    if sort_by == "name":
        finance_data.sort(key=lambda x: x["learner_name"])
    elif sort_by == "mobile":
        finance_data.sort(key=lambda x: x["mobile_no"] or "")
    elif sort_by == "course":
        finance_data.sort(key=lambda x: x["course"])
    elif sort_by == "batch":
        finance_data.sort(key=lambda x: x["batch"])
    elif sort_by == "total_paid":
        finance_data.sort(key=lambda x: float(x["total_paid"] or 0), reverse=True)
    elif sort_by == "balance":
        finance_data.sort(key=lambda x: float(x["balance_fees"] or 0), reverse=True)
    elif sort_by == "profit":
        finance_data.sort(key=lambda x: float(x["profit"] or 0), reverse=True)

    context = {
        "finance_data": finance_data,
        "total_profit": total_profit,
        "total_learner_paid": total_learner_paid,
        "total_mkcl_paid": total_mkcl_paid,
        "selected_year": selected_year,
        "available_years": available_years,
        "available_courses": available_courses,
        "available_batches": formatted_batches,
        "search_query": search_query,
        "course_filter": course_filter,
        "batch_filter": batch_filter,
        "sort_by": sort_by,
        "active_page": "student_finance_details",
    }

    return render(request, "core/student_finance_details.html", context)


@login_required
@staff_member_required
@require_http_methods(["POST"])
@csrf_protect
def update_finance_detail(request):
    """AJAX endpoint to update finance details"""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Invalid JSON"}, status=400)

    student_id = data.get("student_id")
    field = data.get("field")
    raw_value = data.get("value", "0")

    try:
        value = Decimal(str(raw_value)) if raw_value not in ("", None) else Decimal("0.00")
    except (InvalidOperation, TypeError, ValueError):
        return JsonResponse({"success": False, "error": "Invalid amount"}, status=400)

    if value < 0:
        return JsonResponse({"success": False, "error": "Amount cannot be negative"}, status=400)

    try:
        student = AdmittedStudent.objects.get(id=student_id)
    except AdmittedStudent.DoesNotExist:
        return JsonResponse({"success": False, "error": "Student not found"}, status=404)

    finance_detail, _ = StudentFinanceDetail.objects.get_or_create(student=student)

    # Only allow updates to MKCL fees; learner fees come from actual payments.
    if field == "mkcl_1":
        finance_detail.fees_paid_to_mkcl_1 = value
    elif field == "mkcl_2":
        finance_detail.fees_paid_to_mkcl_2 = value
    elif field == "mkcl_3":
        finance_detail.fees_paid_to_mkcl_3 = value
    else:
        return JsonResponse(
            {"success": False, "error": "Cannot update learner fees. These are based on actual payment records."},
            status=400,
        )

    finance_detail.save()

    total_paid = student.paid_fees or Decimal("0.00")
    mkcl_total = (
        (finance_detail.fees_paid_to_mkcl_1 or Decimal("0.00"))
        + (finance_detail.fees_paid_to_mkcl_2 or Decimal("0.00"))
        + (finance_detail.fees_paid_to_mkcl_3 or Decimal("0.00"))
    )
    profit = total_paid - mkcl_total

    return JsonResponse({"success": True, "mkcl_total": float(mkcl_total), "profit": float(profit)})


@login_required
def month_wise_admission(request):
    """Month wise admission details view"""
    from datetime import datetime

    current_year = datetime.now().year
    selected_year = request.GET.get("year", str(current_year))  # Default to current year
    try:
        selected_year_int = int(selected_year)
    except (TypeError, ValueError):
        selected_year_int = current_year
        selected_year = str(current_year)

    # Get all years for filter
    years = AdmittedStudent.objects.dates("admission_date", "year", order="DESC")
    available_years = [date.year for date in years]

    # Get admitted students
    students = AdmittedStudent.objects.all()
    if selected_year:
        students = students.filter(admission_date__year=selected_year_int)

    student_rows = list(students)

    def get_display_course_name(student):
        if student.course == "Other" and student.custom_course:
            return student.custom_course
        return student.course or ""

    all_courses = sorted(
        {get_display_course_name(student) for student in student_rows if get_display_course_name(student)}
    )

    months = ["jan", "feb", "march", "april", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]

    # Initialize data structure
    monthly_admission_data = []
    monthly_totals = {month: 0 for month in months}
    grand_total = 0

    students_by_course_and_month = {}
    for student in student_rows:
        display_course = get_display_course_name(student)
        if not display_course or not student.admission_date:
            continue
        students_by_course_and_month.setdefault((display_course, student.admission_date.month), []).append(student)

    # Count admissions by course and month
    for course in all_courses:
        course_data = {"course": course}
        course_total = 0

        for month_num, month_key in enumerate(months, 1):
            count = len(students_by_course_and_month.get((course, month_num), []))

            course_data[month_key] = count if count > 0 else "-"
            if count > 0:
                monthly_totals[month_key] += count
                course_total += count

        course_data["total"] = course_total if course_total > 0 else "-"
        monthly_admission_data.append(course_data)
        grand_total += course_total

    # Convert monthly_totals zeros to '-'
    for month_key in months:
        if monthly_totals[month_key] == 0:
            monthly_totals[month_key] = "-"

    # Calculate monthly profit data by course using StudentFinanceDetail
    # This takes the total profit from each student's finance detail and adds it up by course and admission month
    monthly_profit_data = []
    profit_monthly_totals = {month: Decimal("0.00") for month in months}
    profit_grand_total = Decimal("0.00")

    for course in all_courses:
        course_profit = {"course": course}
        course_profit_total = Decimal("0.00")

        for month_num, month_key in enumerate(months, 1):
            course_students = students_by_course_and_month.get((course, month_num), [])

            # Sum the profit from StudentFinanceDetail for these students
            month_profit = Decimal("0.00")

            for student in course_students:
                # Get the finance detail for this student
                try:
                    finance_detail = StudentFinanceDetail.objects.get(student=student)
                    # Use the profit property which calculates: paid_fees - total_mkcl_fees
                    student_profit = Decimal(str(finance_detail.profit or 0))
                    month_profit += student_profit
                except StudentFinanceDetail.DoesNotExist:
                    # If no finance detail exists, calculate it manually
                    pass

            # Format and store
            if month_profit != 0:
                course_profit[month_key] = f"₹ {month_profit:.2f}"
                profit_monthly_totals[month_key] += month_profit
                course_profit_total += month_profit
            else:
                course_profit[month_key] = "-"

        course_profit["total"] = f"₹ {course_profit_total:.2f}" if course_profit_total != 0 else "-"
        monthly_profit_data.append(course_profit)
        profit_grand_total += course_profit_total

    # Format monthly profit totals
    monthly_profit_totals_formatted = {}
    for month_key in months:
        if profit_monthly_totals[month_key] != 0:
            monthly_profit_totals_formatted[month_key] = f"₹ {profit_monthly_totals[month_key]:.2f}"
        else:
            monthly_profit_totals_formatted[month_key] = "-"

    context = {
        "monthly_admission_data": monthly_admission_data,
        "monthly_totals": monthly_totals,
        "grand_total": grand_total if grand_total > 0 else "-",
        "monthly_profit_data": monthly_profit_data,
        "monthly_profit_totals": monthly_profit_totals_formatted,
        "profit_grand_total": f"₹ {profit_grand_total:.2f}" if profit_grand_total > 0 else "₹ 0.00",
        "selected_year": selected_year or str(current_year),
        "available_years": available_years,
        "active_page": "month_wise_admission",
    }

    return render(request, "core/month_wise_admission.html", context)


# ================= PAYMENT TRACKING PAGE =================
@login_required(login_url="login")
def payment_tracking(request):
    """Show students whose 1st installment was paid before X days"""
    from datetime import date, timedelta

    # Get number of days from request, default to 25
    days = request.GET.get("days", "25")
    try:
        days = int(days)
        if days < 1:
            days = 25
    except (ValueError, TypeError):
        days = 25

    # Get date X days ago
    cutoff_date = date.today() - timedelta(days=days)

    # Get all students who have at least one payment
    students_with_payments = AdmittedStudent.objects.filter(fee_payments__isnull=False).distinct()

    # Filter students where the earliest (1st) payment was before 25 days
    eligible_students = []
    for student in students_with_payments:
        all_payments = student.fee_payments.all()
        if not all_payments.exists():
            continue

        first_payment = all_payments.order_by("payment_date").first()
        if not first_payment or not first_payment.payment_date:
            continue

        if first_payment.payment_date <= cutoff_date:
            # Get payment summary
            total_paid = all_payments.aggregate(Sum("amount"))["amount__sum"] or Decimal("0")
            total_fees = student.total_fees or Decimal("0")
            remaining = total_fees - total_paid

            # Only include students with remaining fees to pay
            if remaining > 0:
                last_payment = all_payments.order_by("-payment_date").first()
                last_payment_date = last_payment.payment_date if last_payment and last_payment.payment_date else None

                eligible_students.append(
                    {
                        "student": student,
                        "first_payment_date": first_payment.payment_date,
                        "total_paid": total_paid,
                        "total_fees": total_fees,
                        "remaining": remaining,
                        "payment_count": all_payments.count(),
                        "last_payment_date": last_payment_date,
                    }
                )

    # Sort by first payment date
    eligible_students = sorted(eligible_students, key=lambda x: x["first_payment_date"])

    # Pagination
    paginator = Paginator(eligible_students, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "eligible_students": page_obj.object_list,
        "total_count": len(eligible_students),
        "days": days,
        "cutoff_date": cutoff_date,
        "active_page": "payment_tracking",
    }

    return render(request, "core/payment_tracking.html", context)


@login_required(login_url="login")
def payment_tracking_student_detail(request, student_id):
    """Get student details for modal display in payment tracking"""
    student = get_object_or_404(AdmittedStudent, id=student_id)

    # Get payment summary
    all_payments = student.fee_payments.all()
    total_paid = all_payments.aggregate(Sum("amount"))["amount__sum"] or Decimal("0")
    total_fees = student.total_fees or Decimal("0")
    remaining = total_fees - total_paid

    # Get payment history
    payments = all_payments.order_by("-payment_date").values("receipt_no", "amount", "payment_date", "payment_mode")

    data = {
        "id": student.id,
        "full_name": student.full_name,
        "student_name": student.student_name,
        "father_name": student.father_name,
        "mother_name": student.mother_name,
        "date_of_birth": str(student.date_of_birth) if student.date_of_birth else "",
        "gender": student.gender,
        "marital_status": student.marital_status,
        "mobile_own": student.mobile_own,
        "parent_mobile": student.parent_mobile,
        "address": student.address or "",
        "city": student.city or "",
        "tehsil_block": student.tehsil_block or "",
        "district": student.district or "",
        "pin_code": student.pin_code or "",
        "educational_qualification": student.educational_qualification or "",
        "course": student.course or "",
        "batch_month": student.batch_month or "",
        "batch_year": student.batch_year or "",
        "admission_date": str(student.admission_date) if student.admission_date else "",
        "photo": student.photo.url if student.photo else "",
        "total_fees": str(total_fees),
        "total_paid": str(total_paid),
        "remaining": str(remaining),
        "payment_count": all_payments.count(),
        "payments": list(payments),
    }

    return JsonResponse(data)


# ================= ATTENDANCE MANAGEMENT SYSTEM =================


@login_required
def attendance_management(request):
    """Attendance management has been removed"""
    from django.http import HttpResponse

    return HttpResponse("Attendance management feature has been removed from this application.", status=410)


@login_required
def get_attendance_stats(request):
    """Feature removed"""
    return JsonResponse({"error": "Attendance management feature has been removed"}, status=410)


@login_required
def get_daily_attendance_students(request):
    """Feature removed"""
    return JsonResponse({"error": "Attendance management feature has been removed"}, status=410)


@login_required
def save_daily_attendance(request):
    """Feature removed"""
    return JsonResponse({"error": "Attendance management feature has been removed"}, status=410)


@login_required
def get_monthly_timetable(request):
    """Feature removed"""
    return JsonResponse({"error": "Attendance management feature has been removed"}, status=410)


@login_required
def get_monthly_report(request):
    """Feature removed"""
    return JsonResponse({"error": "Attendance management feature has been removed"}, status=410)


@login_required
def export_attendance_excel(request):
    """Feature removed"""
    return JsonResponse({"error": "Attendance management feature has been removed"}, status=410)


@login_required
def export_attendance_pdf(request):
    """Feature removed"""
    return JsonResponse({"error": "Attendance management feature has been removed"}, status=410)


# ============= TIMETABLE & ATTENDANCE MANAGEMENT =============


@login_required
@staff_member_required
def student_timetable(request):
    """Display student timetable with batch assignments"""
    search_query = request.GET.get("search", "").strip()
    course_filter = request.GET.get("course", "")
    batch_month_filter = request.GET.get("batch_month", "")
    batch_year_filter = request.GET.get("batch_year", "")
    theory_batch_filter = request.GET.get("theory_batch", "")
    practical_batch_filter = request.GET.get("practical_batch", "")
    gender_filter = request.GET.get("gender", "")

    # Get all students with batch assignments
    students = AdmittedStudent.objects.all()

    # Apply search filter
    if search_query:
        students = students.filter(
            Q(full_name__icontains=search_query)
            | Q(student_name__icontains=search_query)
            | Q(mobile_own__icontains=search_query)
        )

    # Apply course filter
    if course_filter:
        students = students.filter(course=course_filter)

    # Apply batch month filter
    if batch_month_filter:
        students = students.filter(batch_month=batch_month_filter)

    # Apply batch year filter
    if batch_year_filter:
        students = students.filter(batch_year=batch_year_filter)

    # Apply theory batch filter
    if theory_batch_filter:
        students = students.filter(theory_batch_time=theory_batch_filter)

    # Apply practical batch filter
    if practical_batch_filter:
        students = students.filter(practical_batch_time=practical_batch_filter)

    # Apply gender filter
    if gender_filter:
        students = students.filter(gender=gender_filter)

    # Paginate results
    paginator = Paginator(students, 25)
    page_number = request.GET.get("page", 1)
    page_obj = paginator.get_page(page_number)

    # Get available courses from AdmittedStudent
    course_choices = [
        ("MS-CIT", "MS-CIT"),
        ("Tally", "Tally"),
        ("Advance Excel", "Advance Excel"),
        ("IOT", "IOT"),
        ("Scratch", "Scratch"),
        ("Other", "Other"),
    ]
    all_courses = course_choices

    # Get available batch months and years
    available_batch_months = (
        AdmittedStudent.objects.exclude(batch_month__isnull=True)
        .exclude(batch_month="")
        .values_list("batch_month", flat=True)
        .distinct()
        .order_by("batch_month")
    )

    available_batch_years = (
        AdmittedStudent.objects.exclude(batch_year__isnull=True)
        .exclude(batch_year="")
        .values_list("batch_year", flat=True)
        .distinct()
        .order_by("-batch_year")
    )

    # Get available batches
    available_theory_batches = (
        Batch.objects.filter(batch_type="Theory", course__isnull=True)
        .values_list("time_slot", flat=True)
        .distinct()
        .order_by("time_slot")
    )

    available_practical_batches = (
        Batch.objects.filter(batch_type="Practical", course__isnull=True)
        .values_list("time_slot", flat=True)
        .distinct()
        .order_by("time_slot")
    )

    # Calculate total unique available time slots
    all_available_slots = set(list(available_theory_batches) + list(available_practical_batches))
    total_available_slots = len(all_available_slots)

    time_slot_display_map = TIME_SLOT_DISPLAY_MAP
    time_slots = get_cached_time_slots()

    context = {
        "page_obj": page_obj,
        "students": page_obj.object_list,
        "search_query": search_query,
        "course_filter": course_filter,
        "batch_month_filter": batch_month_filter,
        "batch_year_filter": batch_year_filter,
        "theory_batch_filter": theory_batch_filter,
        "practical_batch_filter": practical_batch_filter,
        "gender_filter": gender_filter,
        "all_courses": all_courses,
        "available_batch_months": available_batch_months,
        "available_batch_years": available_batch_years,
        "available_theory_batches": available_theory_batches,
        "available_practical_batches": available_practical_batches,
        "total_available_slots": total_available_slots,
        "time_slot_display_map": time_slot_display_map,
        "total_students": students.count(),
        "time_slots": time_slots,
        "active_page": "student_timetable",
    }

    return render(request, "core/timetable/student_timetable.html", context)


@login_required
@staff_member_required
def edit_student_batch(request, student_id):
    """Edit student's theory and practical batch assignments"""
    student = get_object_or_404(AdmittedStudent, id=student_id)

    if request.method == "POST":
        theory_batch = request.POST.get("theory_batch_time", "").strip()
        practical_batch = request.POST.get("practical_batch_time", "").strip()

        # Update student batches
        if theory_batch:
            student.theory_batch_time = theory_batch
        if practical_batch:
            student.practical_batch_time = practical_batch

        student.save()
        messages.success(request, f"✅ Batch timing updated for {student.full_name}")
        return redirect("student_timetable")

    # Get only existing theory and practical batches
    theory_batches = Batch.objects.filter(batch_type="Theory", course__isnull=True).values_list("time_slot", flat=True)
    practical_batches = Batch.objects.filter(batch_type="Practical", course__isnull=True).values_list(
        "time_slot", flat=True
    )

    time_slots = TIME_SLOT_CHOICES

    # Filter to only show existing batches
    available_theory_slots = [(slot, display) for slot, display in time_slots if slot in theory_batches]
    available_practical_slots = [(slot, display) for slot, display in time_slots if slot in practical_batches]

    context = {
        "student": student,
        "theory_batches": available_theory_slots,
        "practical_batches": available_practical_slots,
        "active_page": "student_timetable",
    }

    return render(request, "core/timetable/edit_batch.html", context)


@login_required
@staff_member_required
def batch_overview_dashboard(request):
    """Dashboard showing batch-wise student distribution"""
    time_slots = TIME_SLOT_CHOICES

    # Get only existing theory batches from database
    theory_batches = []
    for batch in Batch.objects.filter(batch_type="Theory", course__isnull=True):
        count = AdmittedStudent.objects.filter(theory_batch_time=batch.time_slot).count()
        display = dict(time_slots).get(batch.time_slot, batch.time_slot)
        theory_batches.append(
            {"id": batch.id, "slot": batch.time_slot, "display": display, "count": count, "capacity": batch.capacity}
        )

    # Get only existing practical batches from database
    practical_batches = []
    for batch in Batch.objects.filter(batch_type="Practical", course__isnull=True):
        count = AdmittedStudent.objects.filter(practical_batch_time=batch.time_slot).count()
        display = dict(time_slots).get(batch.time_slot, batch.time_slot)
        practical_batches.append(
            {"id": batch.id, "slot": batch.time_slot, "display": display, "count": count, "capacity": batch.capacity}
        )

    # Total statistics
    total_students = AdmittedStudent.objects.count()
    students_with_theory = (
        AdmittedStudent.objects.filter(theory_batch_time__isnull=False).exclude(theory_batch_time="").count()
    )
    students_with_practical = (
        AdmittedStudent.objects.filter(practical_batch_time__isnull=False).exclude(practical_batch_time="").count()
    )

    context = {
        "theory_batches": theory_batches,
        "practical_batches": practical_batches,
        "total_students": total_students,
        "students_with_theory": students_with_theory,
        "students_with_practical": students_with_practical,
        "students_without_batch": total_students - max(students_with_theory, students_with_practical),
        "active_page": "batch_overview",
    }

    return render(request, "core/timetable/batch_overview.html", context)


@login_required
@staff_member_required
def mark_attendance_page(request):
    """Page to mark attendance for a specific batch and date"""
    from datetime import date as date_class

    if request.method == "POST":
        attendance_date = request.POST.get("attendance_date")
        batch_time = request.POST.get("batch_time")
        batch_type = request.POST.get("batch_type")

        if not all([attendance_date, batch_time, batch_type]):
            messages.error(request, "❌ Please select date, batch time, and batch type")
            return redirect("mark_attendance")

        return redirect("save_attendance", date=attendance_date, batch_time=batch_time, batch_type=batch_type)

    # Get only existing theory batches
    theory_batches = set(
        Batch.objects.filter(batch_type="Theory", course__isnull=True)
        .order_by()
        .values_list("time_slot", flat=True)
        .distinct()
    )

    # Get only existing practical batches
    practical_batches = set(
        Batch.objects.filter(batch_type="Practical", course__isnull=True)
        .order_by()
        .values_list("time_slot", flat=True)
        .distinct()
    )

    # Create filtered time slots for display
    theory_slots = [(slot, display) for slot, display in TIME_SLOT_CHOICES if slot in theory_batches]
    practical_slots = [(slot, display) for slot, display in TIME_SLOT_CHOICES if slot in practical_batches]

    context = {
        "theory_slots": theory_slots,
        "practical_slots": practical_slots,
        "today": date_class.today().isoformat(),
        "active_page": "mark_attendance",
    }

    return render(request, "core/timetable/mark_attendance.html", context)


@login_required
@staff_member_required
def save_attendance(request, date, batch_time, batch_type):
    """Save attendance for students in a specific batch"""
    from datetime import datetime as dt

    # Parse date
    try:
        attendance_date = dt.strptime(date, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        messages.error(request, "❌ Invalid date format")
        return redirect("mark_attendance")

    # Get students for the selected batch
    if batch_type == "theory":
        students = AdmittedStudent.objects.filter(
            theory_batch_time=batch_time,
            batch_status="active",
            is_archived=False,
        ).order_by("full_name")
    else:  # practical
        students = AdmittedStudent.objects.filter(
            practical_batch_time=batch_time,
            batch_status="active",
            is_archived=False,
        ).order_by("full_name")

    if request.method == "POST":
        with transaction.atomic():
            for student in students:
                # Get attendance status
                attendance_key = f"attendance_{student.id}"
                status = request.POST.get(attendance_key, "A")
                remarks = request.POST.get(f"remarks_{student.id}", "").strip()

                # Create or update attendance record
                attendance, created = Attendance.objects.get_or_create(
                    student=student,
                    date=attendance_date,
                    defaults={
                        "marked_by": request.user,
                        "theory_attendance": "",
                        "practical_attendance": "",
                    },
                )

                # Update attendance based on batch type
                if batch_type == "theory":
                    attendance.theory_attendance = status
                    if created:
                        attendance.practical_attendance = ""
                else:
                    attendance.practical_attendance = status
                    if created:
                        attendance.theory_attendance = ""

                if remarks:
                    attendance.remarks = remarks

                attendance.marked_by = request.user
                attendance.save()

        messages.success(request, f"✅ Attendance saved for {batch_type.capitalize()} batch on {attendance_date}")
        return redirect("attendance_reports")

    time_slot_display_map = TIME_SLOT_DISPLAY_MAP

    # Get time slot display
    time_slot_display = time_slot_display_map.get(batch_time, batch_time)

    context = {
        "students": students,
        "attendance_date": attendance_date,
        "batch_time": batch_time,
        "time_slot_display": time_slot_display,
        "batch_type": batch_type,
        "status_choices": [("P", "Present"), ("A", "Absent"), ("L", "Leave"), ("H", "Holiday")],
        "active_page": "mark_attendance",
    }

    return render(request, "core/timetable/save_attendance.html", context)


@login_required
@staff_member_required
def attendance_reports(request):
    """Comprehensive attendance reports"""
    from datetime import date as date_class

    report_type = request.GET.get("report_type", "student")
    if report_type == "daily":
        return redirect("attendance_reports")

    # Student Attendance Report
    if report_type == "student":
        students = AdmittedStudent.objects.prefetch_related(
            Prefetch("attendance_records", queryset=Attendance.objects.order_by("-date"))
        ).order_by("full_name")
        attendance_dates = list(Attendance.objects.order_by("date").values_list("date", flat=True).distinct())

        student_reports = []
        for student in students:
            attendance_records = list(student.attendance_records.all())
            total_records = len(attendance_records)
            theory_present_count = sum(1 for record in attendance_records if record.theory_attendance == "P")
            theory_absent_count = sum(1 for record in attendance_records if record.theory_attendance == "A")
            practical_present_count = sum(1 for record in attendance_records if record.practical_attendance == "P")
            practical_absent_count = sum(1 for record in attendance_records if record.practical_attendance == "A")
            present_count = sum(
                1
                for record in attendance_records
                if record.theory_attendance == "P" or record.practical_attendance == "P"
            )
            absent_count = sum(
                1
                for record in attendance_records
                if record.theory_attendance == "A" and record.practical_attendance == "A"
            )

            if total_records > 0:
                percentage = (present_count / total_records) * 100
            else:
                percentage = 0

            attendance_by_date = {record.date: record for record in attendance_records}
            theory_days = []
            practical_days = []

            for attendance_date in attendance_dates:
                record = attendance_by_date.get(attendance_date)
                theory_status = ""
                practical_status = ""

                if record:
                    theory_status = record.theory_attendance or ""
                    practical_status = record.practical_attendance or ""

                theory_days.append(
                    {
                        "date": attendance_date,
                        "status": theory_status,
                    }
                )
                practical_days.append(
                    {
                        "date": attendance_date,
                        "status": practical_status,
                    }
                )

            student_reports.append(
                {
                    "student": student,
                    "total": total_records,
                    "present": present_count,
                    "absent": absent_count,
                    "theory_present": theory_present_count,
                    "theory_absent": theory_absent_count,
                    "practical_present": practical_present_count,
                    "practical_absent": practical_absent_count,
                    "records": attendance_records,
                    "theory_days": theory_days,
                    "practical_days": practical_days,
                    "percentage": round(percentage, 2),
                }
            )

        context = {
            "report_type": "student",
            "student_reports": student_reports,
            "attendance_dates": attendance_dates,
            "title": "Student Attendance Report",
            "active_page": "attendance_reports",
        }

    # Batch Attendance Report
    else:  # batch
        selected_date = request.GET.get("date")

        if selected_date:
            try:
                report_date = dt.strptime(selected_date, "%Y-%m-%d").date()
            except (ValueError, TypeError):
                report_date = date_class.today()
        else:
            report_date = date_class.today()

        time_slot_display_map = TIME_SLOT_DISPLAY_MAP

        # Get all existing theory and practical batch time slots
        theory_slots = (
            Batch.objects.filter(batch_type="Theory", course__isnull=True)
            .values_list("time_slot", flat=True)
            .distinct()
        )

        practical_slots = (
            Batch.objects.filter(batch_type="Practical", course__isnull=True)
            .values_list("time_slot", flat=True)
            .distinct()
        )

        # Combine all unique slots
        all_slots = sorted(set(theory_slots) | set(practical_slots))

        attendance_records = Attendance.objects.filter(date=report_date).select_related("student", "marked_by")
        attendance_by_student_id = {record.student_id: record for record in attendance_records}

        def serialize_batch_students(students, attendance_type):
            serialized = []
            for student in students.order_by("full_name"):
                attendance_record = attendance_by_student_id.get(student.id)
                status = "A"
                if attendance_record:
                    if attendance_type == "theory":
                        status = attendance_record.theory_attendance or "A"
                    else:
                        status = attendance_record.practical_attendance or "A"

                serialized.append(
                    {
                        "id": student.id,
                        "full_name": student.full_name,
                        "mobile_own": student.mobile_own,
                        "status": status,
                    }
                )
            return serialized

        batch_reports = []
        for slot in all_slots:
            display = time_slot_display_map.get(slot, slot)

            # Theory batch
            theory_students = AdmittedStudent.objects.filter(theory_batch_time=slot)
            theory_present = Attendance.objects.filter(
                student__in=theory_students, date=report_date, theory_attendance="P"
            ).count()
            theory_absent = theory_students.count() - theory_present

            # Practical batch
            practical_students = AdmittedStudent.objects.filter(practical_batch_time=slot)
            practical_present = Attendance.objects.filter(
                student__in=practical_students, date=report_date, practical_attendance="P"
            ).count()
            practical_absent = practical_students.count() - practical_present

            batch_reports.append(
                {
                    "slot": slot,
                    "display": display,
                    "theory": {
                        "total": theory_students.count(),
                        "present": theory_present,
                        "absent": theory_absent,
                        "students": serialize_batch_students(theory_students, "theory"),
                    },
                    "practical": {
                        "total": practical_students.count(),
                        "present": practical_present,
                        "absent": practical_absent,
                        "students": serialize_batch_students(practical_students, "practical"),
                    },
                }
            )

        context = {
            "report_type": "batch",
            "report_date": report_date,
            "batch_reports": batch_reports,
            "title": f"Batch Attendance Report - {report_date}",
            "active_page": "attendance_reports",
        }

    return render(request, "core/timetable/attendance_reports.html", context)


@login_required
@staff_member_required
def export_timetable_excel(request):
    """Export student timetable to Excel"""
    students = AdmittedStudent.objects.all().order_by("full_name")

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Timetable"

    # Headers
    headers = ["Student Name", "Gender", "Course", "Theory Batch", "Practical Batch", "Admission Date"]
    ws.append(headers)

    # Style header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add data
    for student in students:
        theory_display = TIME_SLOT_DISPLAY_MAP.get(
            student.theory_batch_time, student.theory_batch_time or "Not Assigned"
        )

        practical_display = TIME_SLOT_DISPLAY_MAP.get(
            student.practical_batch_time, student.practical_batch_time or "Not Assigned"
        )

        ws.append(
            [
                student.full_name,
                student.gender,
                student.course or "N/A",
                theory_display,
                practical_display,
                student.admission_date.strftime("%Y-%m-%d"),
            ]
        )

    # Adjust column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 15

    # Return as attachment
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="student_timetable.xlsx"'
    wb.save(response)
    return response


@login_required
@staff_member_required
def export_attendance_report_excel(request):
    """Export attendance reports to Excel in the same format as the reports page."""
    wb = openpyxl.Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    def style_sheet(sheet, widths):
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        sheet.freeze_panes = "A2"

    student_sheet = wb.active
    student_sheet.title = "Student Attendance"
    attendance_dates = list(Attendance.objects.order_by("date").values_list("date", flat=True).distinct())
    first_header = ["Sr. No", "Student Name", "Mobile No"]
    second_header = ["", "", ""]
    for attendance_date in attendance_dates:
        first_header.extend([attendance_date.strftime("%d-%b"), ""])
        second_header.extend(["Theory", "Practical"])
    student_sheet.append(first_header)
    student_sheet.append(second_header)
    student_sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    student_sheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    student_sheet.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)
    for date_index in range(len(attendance_dates)):
        start_column = 4 + (date_index * 2)
        student_sheet.merge_cells(start_row=1, start_column=start_column, end_row=1, end_column=start_column + 1)

    students = AdmittedStudent.objects.filter(is_archived=False).prefetch_related(
        Prefetch("attendance_records", queryset=Attendance.objects.order_by("date"))
    ).order_by("full_name")
    for index, student in enumerate(students, start=1):
        attendance_records = list(student.attendance_records.all())
        attendance_by_date = {record.date: record for record in attendance_records}
        row = [index, student.full_name, student.mobile_own or ""]
        for attendance_date in attendance_dates:
            record = attendance_by_date.get(attendance_date)
            row.extend(
                [
                    record.theory_attendance if record else "",
                    record.practical_attendance if record else "",
                ]
            )
        student_sheet.append(row)

    for row in student_sheet.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    student_sheet.freeze_panes = "D3"
    student_sheet.column_dimensions["A"].width = 10
    student_sheet.column_dimensions["B"].width = 30
    student_sheet.column_dimensions["C"].width = 16
    for column_index in range(4, student_sheet.max_column + 1):
        student_sheet.column_dimensions[get_column_letter(column_index)].width = 12

    daily_sheet = wb.create_sheet("Daily Attendance")
    daily_sheet.append(["Sr. No", "Date", "Student Name", "Mobile No", "Course", "Theory", "Practical", "Remarks"])

    attendance_records = Attendance.objects.select_related("student").order_by("date", "student__full_name")
    for index, record in enumerate(attendance_records, start=1):
        student = record.student
        course_name = student.custom_course if student.course == "Other" and student.custom_course else student.course
        daily_sheet.append(
            [
                index,
                record.date.strftime("%Y-%m-%d") if record.date else "",
                student.full_name,
                student.mobile_own or "",
                course_name or "N/A",
                record.get_theory_attendance_display(),
                record.get_practical_attendance_display(),
                record.remarks or "",
            ]
        )

    style_sheet(
        daily_sheet,
        {"A": 10, "B": 14, "C": 30, "D": 16, "E": 20, "F": 14, "G": 14, "H": 30},
    )

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="attendance_report.xlsx"'
    wb.save(response)
    return response


# ================= BATCH MANAGEMENT VIEWS =================


@login_required
@staff_member_required
@require_http_methods(["POST"])
@csrf_protect
def create_batch(request):
    """Create a new batch - AJAX endpoint"""
    try:
        data = json.loads(request.body)
        batch_type = data.get("batch_type", "").strip()
        time_slot = data.get("time_slot", "").strip()
        capacity = data.get("capacity", 50)

        # Validate inputs
        if not batch_type or not time_slot:
            return JsonResponse({"success": False, "error": "Batch type and time slot are required"}, status=400)

        if batch_type not in ["Theory", "Practical"]:
            return JsonResponse(
                {"success": False, "error": "Invalid batch type. Must be Theory or Practical"}, status=400
            )

        # Validate time slot
        valid_slots = TIME_SLOT_VALUES
        if time_slot not in valid_slots:
            return JsonResponse({"success": False, "error": "Invalid time slot"}, status=400)

        # Convert capacity to int
        try:
            capacity = int(capacity) if capacity else 50
            if capacity < 1:
                capacity = 50
        except (ValueError, TypeError):
            capacity = 50

        # Check if batch already exists
        from .models import Batch

        existing_batch = Batch.objects.filter(batch_type=batch_type, time_slot=time_slot, course__isnull=True).first()

        if existing_batch:
            return JsonResponse(
                {"success": False, "error": f"{batch_type} batch at {time_slot} already exists"}, status=400
            )

        # Create the batch
        batch = Batch.objects.create(batch_type=batch_type, time_slot=time_slot, capacity=capacity, course=None)

        return JsonResponse(
            {
                "success": True,
                "message": f"✅ New {batch_type} batch created at {batch.get_time_slot_display()}",
                "batch": {
                    "id": batch.id,
                    "type": batch_type,
                    "time_slot": time_slot,
                    "display": batch.get_time_slot_display(),
                    "capacity": capacity,
                },
            },
            status=201,
        )

    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Invalid JSON data"}, status=400)
    except Exception as e:
        print(f"Error creating batch: {str(e)}")
        import traceback

        traceback.print_exc()
        return JsonResponse({"success": False, "error": f"Server error: {str(e)}"}, status=500)


@login_required
@staff_member_required
@require_http_methods(["POST"])
@csrf_protect
def delete_batch(request, batch_id):
    """Delete a batch - AJAX endpoint"""
    try:
        from .models import Batch

        batch = get_object_or_404(Batch, id=batch_id)
        batch_type = batch.batch_type
        time_slot_display = batch.get_time_slot_display()

        # Unassign all students from this batch before deleting
        if batch_type == "Theory":
            students = AdmittedStudent.objects.filter(theory_batch_time=batch.time_slot)
            student_count = students.count()
            # Unassign students
            students.update(theory_batch_time=None)
        else:
            students = AdmittedStudent.objects.filter(practical_batch_time=batch.time_slot)
            student_count = students.count()
            # Unassign students
            students.update(practical_batch_time=None)

        # Delete the batch
        batch.delete()

        message = f"✅ {batch_type} batch at {time_slot_display} has been deleted"
        if student_count > 0:
            message += f" ({student_count} student(s) were unassigned)"

        return JsonResponse({"success": True, "message": message, "students_unassigned": student_count})

    except Batch.DoesNotExist:
        return JsonResponse({"success": False, "error": "Batch not found"}, status=404)
    except Exception as e:
        print(f"Error deleting batch: {str(e)}")
        import traceback

        traceback.print_exc()
        return JsonResponse({"success": False, "error": f"Server error: {str(e)}"}, status=500)


@login_required
@staff_member_required
def get_batch_list(request):
    """Get updated list of theory and practical batches - AJAX endpoint"""
    try:
        from .models import Batch

        time_slots = TIME_SLOT_CHOICES
        active_students = _active_batch_students()

        # Get theory batches that actually exist in database
        theory_batches = []
        for batch in Batch.objects.filter(batch_type="Theory", course__isnull=True, is_archived=False):
            count = active_students.filter(theory_batch_time=batch.time_slot).count()
            display = dict(time_slots).get(batch.time_slot, batch.time_slot)
            theory_batches.append(
                {
                    "id": batch.id,
                    "slot": batch.time_slot,
                    "display": display,
                    "count": count,
                    "capacity": batch.capacity,
                    "exists": True,
                }
            )

        # Get practical batches that actually exist in database
        practical_batches = []
        for batch in Batch.objects.filter(batch_type="Practical", course__isnull=True, is_archived=False):
            count = active_students.filter(practical_batch_time=batch.time_slot).count()
            display = dict(time_slots).get(batch.time_slot, batch.time_slot)
            practical_batches.append(
                {
                    "id": batch.id,
                    "slot": batch.time_slot,
                    "display": display,
                    "count": count,
                    "capacity": batch.capacity,
                    "exists": True,
                }
            )

        # Total statistics
        total_students = active_students.count()
        students_with_theory = (
            active_students.filter(theory_batch_time__isnull=False).exclude(theory_batch_time="").count()
        )
        students_with_practical = (
            active_students.filter(practical_batch_time__isnull=False).exclude(practical_batch_time="").count()
        )

        return JsonResponse(
            {
                "success": True,
                "theory_batches": theory_batches,
                "practical_batches": practical_batches,
                "total_students": total_students,
                "students_with_theory": students_with_theory,
                "students_with_practical": students_with_practical,
                "students_without_batch": total_students - max(students_with_theory, students_with_practical),
            }
        )

    except Exception as e:
        print(f"Error getting batch list: {str(e)}")
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@staff_member_required
def get_batch_id(request):
    """Helper endpoint to get batch ID by type and time_slot"""
    try:
        batch_type = request.GET.get("type")
        time_slot = request.GET.get("time_slot")

        if not batch_type or not time_slot:
            return JsonResponse({"success": False, "error": "Missing type or time_slot parameter"}, status=400)

        # Find the batch - match the same filter used in create_batch
        batch = Batch.objects.get(batch_type=batch_type, time_slot=time_slot, course__isnull=True)

        return JsonResponse({"success": True, "batch_id": batch.id})

    except Batch.DoesNotExist:
        return JsonResponse({"success": False, "error": "Batch not found"}, status=404)
    except Exception as e:
        print(f"Error getting batch ID: {str(e)}")
        import traceback

        traceback.print_exc()
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@login_required
@staff_member_required
@require_http_methods(["POST"])
@csrf_protect
def edit_batch(request, batch_id):
    """Edit a batch - AJAX endpoint"""
    try:
        import json

        from .models import Batch

        batch = get_object_or_404(Batch, id=batch_id)
        data = json.loads(request.body)

        new_time_slot = data.get("time_slot")
        new_capacity = data.get("capacity")

        if not new_time_slot:
            return JsonResponse({"success": False, "error": "Time slot is required"}, status=400)

        # Validate time slot
        valid_slots = TIME_SLOT_VALUES
        if new_time_slot not in valid_slots:
            return JsonResponse({"success": False, "error": "Invalid time slot"}, status=400)

        # Check if another batch already has this time slot
        existing_batch = (
            Batch.objects.filter(batch_type=batch.batch_type, time_slot=new_time_slot, course__isnull=True)
            .exclude(id=batch.id)
            .first()
        )

        if existing_batch:
            return JsonResponse(
                {"success": False, "error": f"A {batch.batch_type.lower()} batch already exists at {new_time_slot}"},
                status=400,
            )

        # Update batch
        batch.time_slot = new_time_slot
        if new_capacity:
            try:
                capacity = int(new_capacity)
                if capacity > 0:
                    batch.capacity = capacity
            except (ValueError, TypeError):
                pass

        batch.save()

        return JsonResponse(
            {
                "success": True,
                "message": f"✅ {batch.batch_type} batch updated successfully",
                "batch": {
                    "id": batch.id,
                    "type": batch.batch_type,
                    "time_slot": batch.time_slot,
                    "capacity": batch.capacity,
                },
            }
        )

    except Batch.DoesNotExist:
        return JsonResponse({"success": False, "error": "Batch not found"}, status=404)
    except Exception as e:
        print(f"Error editing batch: {str(e)}")
        import traceback

        traceback.print_exc()
        return JsonResponse({"success": False, "error": f"Server error: {str(e)}"}, status=500)


@login_required
@staff_member_required
def get_batch_students(request):
    """Get all students in a specific batch"""
    try:
        batch_type = request.GET.get("batch_type", "")
        time_slot = request.GET.get("time_slot", "")

        if not batch_type or not time_slot:
            return JsonResponse({"success": False, "error": "Missing batch_type or time_slot"}, status=400)

        # Batch overview should only show students currently eligible for active batches.
        if batch_type == "Theory":
            students = (
                AdmittedStudent.objects.filter(
                    theory_batch_time=time_slot,
                    batch_status="active",
                    is_archived=False,
                )
                .values("id", "full_name", "gender", "theory_batch_time", "practical_batch_time")
                .order_by("full_name")
            )
        else:  # Practical
            students = (
                AdmittedStudent.objects.filter(
                    practical_batch_time=time_slot,
                    batch_status="active",
                    is_archived=False,
                )
                .values("id", "full_name", "gender", "theory_batch_time", "practical_batch_time")
                .order_by("full_name")
            )

        students_list = list(students)

        return JsonResponse({"success": True, "students": students_list})
    except Exception as e:
        print(f"Error fetching batch students: {str(e)}")
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@login_required
@staff_member_required
@require_http_methods(["POST"])
@csrf_protect
def update_batch_students(request):
    """Update student batch assignments"""
    try:
        data = json.loads(request.body)
        changes = data.get("changes", [])
        _validate_batch_assignment_changes(changes)

        for change in changes:
            student_id = change.get("student_id")
            field_type = change.get("type")
            value = (change.get("value") or "").strip()

            student = AdmittedStudent.objects.get(id=student_id)

            if field_type == "theory_batch_time":
                student.theory_batch_time = value if value else None
            elif field_type == "practical_batch_time":
                student.practical_batch_time = value if value else None

            student.save()

        return JsonResponse({"success": True, "message": f"Updated {len(changes)} student(s)"})
    except ValueError as e:
        return JsonResponse({"success": False, "error": str(e)}, status=400)
    except AdmittedStudent.DoesNotExist:
        return JsonResponse({"success": False, "error": "Student not found"}, status=404)
    except Exception as e:
        print(f"Error updating batch students: {str(e)}")
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@login_required
@staff_member_required
def get_all_students(request):
    """Get all admitted students for adding to batches"""
    try:
        students = (
            _active_batch_students()
            .values("id", "full_name", "gender", "theory_batch_time", "practical_batch_time")
            .order_by("full_name")
        )

        return JsonResponse({"success": True, "students": list(students)})
    except Exception as e:
        print(f"Error fetching students: {str(e)}")
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@login_required
@staff_member_required
def get_student_detail_batch(request, student_id):
    """Get student details for batch modal display"""
    try:
        student = AdmittedStudent.objects.get(id=student_id)
        fee_payments = FeePayment.objects.filter(student=student).order_by("payment_date")
        payment_history = [
            {
                "id": payment.id,
                "payment_date": payment.payment_date.strftime("%d-%m-%Y") if payment.payment_date else "",
                "payment_time": "",
                "amount": float(payment.amount),
                "payment_mode": payment.payment_mode,
                "receipt_no": payment.receipt_no or "",
                "remaining_after": float(payment.remaining_after_this),
            }
            for payment in fee_payments
        ]

        data = {
            "success": True,
            "student": {
                "id": student.id,
                "student_name": student.student_name,
                "father_name": student.father_name,
                "surname": student.surname,
                "mother_name": student.mother_name or "",
                "full_name": student.full_name,
                "gender": student.gender,
                "mobile_own": student.mobile_own,
                "parent_mobile": student.parent_mobile or "",
                "email": getattr(student, "email", "") or "",
                "course": student.course or "",
                "custom_course": student.custom_course or "",
                "educational_qualification": student.educational_qualification or "",
                "batch_month": student.batch_month or "",
                "batch_year": student.batch_year or "",
                "batch_display": student.batch_display or "Not Assigned",
                "batch_status": student.batch_status,
                "theory_batch_time": student.theory_batch_time or "",
                "practical_batch_time": student.practical_batch_time or "",
                "admission_date": student.admission_date.strftime("%Y-%m-%d") if student.admission_date else "",
                "address": student.address or "",
                "city": student.city or "",
                "tehsil_block": student.tehsil_block or "",
                "district": student.district or "",
                "pincode": student.pin_code or "",
                "pin_code": student.pin_code or "",
                "photo": student.photo.url if student.photo else None,
                "total_fees": float(student.total_fees) if student.total_fees else 0,
                "paid_fees": float(student.paid_fees) if student.paid_fees else 0,
                "remaining_fees": float(student.remaining_fees) if student.remaining_fees else 0,
                "payment_history": payment_history,
            },
        }

        return JsonResponse(data)
    except AdmittedStudent.DoesNotExist:
        return JsonResponse({"success": False, "error": "Student not found"}, status=404)
    except Exception as e:
        print(f"Error fetching student details: {str(e)}")
        return JsonResponse({"success": False, "error": str(e)}, status=500)


# ================= UPDATE BATCH CAPACITY =================
@login_required
@staff_member_required
@require_http_methods(["POST"])
@csrf_protect
def update_batch_capacity(request):
    """Update batch capacity"""
    try:
        data = json.loads(request.body)
        batch_id = data.get("batch_id")
        new_capacity = data.get("capacity")

        if not batch_id or not new_capacity:
            return JsonResponse({"success": False, "error": "Batch ID and capacity are required"}, status=400)

        # Get the batch
        batch = Batch.objects.get(id=batch_id)

        # Update capacity
        batch.capacity = int(new_capacity)
        batch.save()

        return JsonResponse(
            {
                "success": True,
                "message": f"Batch capacity updated to {new_capacity}",
                "batch": {"id": batch.id, "capacity": batch.capacity},
            }
        )

    except Batch.DoesNotExist:
        return JsonResponse({"success": False, "error": "Batch not found"}, status=404)
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Invalid JSON"}, status=400)
    except ValueError as e:
        return JsonResponse({"success": False, "error": f"Invalid capacity value: {str(e)}"}, status=400)
    except Exception as e:
        print(f"Error updating batch capacity: {str(e)}")
        return JsonResponse({"success": False, "error": str(e)}, status=500)


NOTIFICATION_PRIORITY_STYLES = {
    "urgent": "danger",
    "pending": "warning",
    "success": "success",
    "info": "info",
}

THREAD_SCOPE_MODEL_MAP = {
    "student": AdmittedStudent,
    "enquiry": Enquiry,
    "finance": FeePayment,
}


def _user_role_label(user):
    """Return a human-readable role label for a user."""
    if user.is_superuser:
        return "Super Admin"
    if user.groups.filter(name="Admin").exists():
        return "Admin"
    if user.groups.filter(name="Counselor").exists():
        return "Counselor"
    if user.groups.filter(name="Accountant").exists():
        return "Accountant"
    if user.groups.filter(name="Attendance Manager").exists():
        return "Attendance Staff"
    return "Team Member"


def _serialize_notification(notification):
    """Serialize a notification for template and API payloads."""
    return {
        "id": notification.id,
        "title": notification.title,
        "message": notification.message,
        "category": notification.category,
        "category_label": notification.get_category_display(),
        "priority": notification.priority,
        "priority_label": notification.get_priority_display(),
        "priority_style": NOTIFICATION_PRIORITY_STYLES.get(notification.priority, "info"),
        "is_read": notification.is_read,
        "action_label": notification.action_label,
        "link_url": notification.link_url,
        "created_at": timezone.localtime(notification.created_at).strftime("%d %b %Y, %I:%M %p"),
        "due_at": timezone.localtime(notification.due_at).strftime("%d %b %Y, %I:%M %p") if notification.due_at else "",
        "actor": notification.actor.username if notification.actor else "System",
    }


def _serialize_thread(thread, user):
    """Serialize a collaboration thread with unread state for a user."""
    latest_entry = thread.entries.select_related("author").order_by("-created_at").first()
    participant_state = thread.participant_states.filter(user=user).first()
    unread = bool(
        latest_entry
        and (
            participant_state is None
            or participant_state.last_read_at is None
            or latest_entry.created_at > participant_state.last_read_at
        )
    )
    return {
        "id": thread.id,
        "title": thread.title,
        "scope": thread.scope,
        "status": thread.status,
        "tags": thread.tags or [],
        "is_pinned": thread.is_pinned,
        "assigned_to": thread.assigned_to.username if thread.assigned_to else "",
        "last_activity_at": timezone.localtime(thread.last_activity_at).strftime("%d %b %Y, %I:%M %p"),
        "preview": (
            (latest_entry.body[:117] + "...")
            if latest_entry and len(latest_entry.body) > 120
            else (latest_entry.body if latest_entry else "")
        ),
        "latest_author": latest_entry.author.username if latest_entry and latest_entry.author else "System",
        "unread": unread,
        "entry_count": thread.entries.count(),
    }


def _serialize_comment(entry):
    """Serialize a thread comment for API responses."""
    return {
        "id": entry.id,
        "thread_id": entry.thread_id,
        "parent_id": entry.parent_id,
        "body": entry.body,
        "author": entry.author.username if entry.author else "System",
        "role_label": _user_role_label(entry.author) if entry.author else "System",
        "attachment_url": entry.attachment.url if entry.attachment else "",
        "status_update": entry.status_update,
        "mentions": list(entry.mentions.values_list("username", flat=True)),
        "created_at": timezone.localtime(entry.created_at).strftime("%d %b %Y, %I:%M %p"),
    }


def _resolve_thread_target(scope, object_id):
    """Resolve a thread scope and object id into a concrete model instance."""
    model = THREAD_SCOPE_MODEL_MAP.get(scope)
    if model is None or not object_id:
        return None
    return get_object_or_404(model, pk=object_id)


@login_required
def notifications_page(request):
    """Render the full notifications center page."""
    category = request.GET.get("category", "").strip()
    status_filter = request.GET.get("status", "").strip()
    notifications = Notification.objects.filter(recipient=request.user).select_related("actor")
    if category:
        notifications = notifications.filter(category=category)
    if status_filter == "unread":
        notifications = notifications.filter(is_read=False)
    elif status_filter == "read":
        notifications = notifications.filter(is_read=True)

    return render(
        request,
        "core/notifications_center.html",
        {
            "active_page": "notifications",
            "notification_items": notifications[:80],
            "notification_category": category,
            "notification_status": status_filter,
            "notification_categories": NotificationSetting.CATEGORY_CHOICES,
        },
    )


@login_required
def notifications_feed_api(request):
    """Return a filtered notification feed for the current user."""
    category = request.GET.get("category", "").strip()
    limit = min(int(request.GET.get("limit", 12) or 12), 50)
    notifications = Notification.objects.filter(recipient=request.user).select_related("actor")
    if category:
        notifications = notifications.filter(category=category)
    unread_count = notifications.filter(is_read=False).count()
    items = [_serialize_notification(item) for item in notifications[:limit]]
    return JsonResponse({"success": True, "unread_count": unread_count, "items": items})


@login_required
@require_http_methods(["POST"])
@csrf_protect
def notification_mark_read(request, notification_id):
    """Toggle read state for a single notification."""
    notification = get_object_or_404(Notification, id=notification_id, recipient=request.user)
    payload = json.loads(request.body or "{}") if request.body else {}
    is_read = bool(payload.get("is_read", True))
    notification.is_read = is_read
    notification.save(update_fields=["is_read", "updated_at"])
    return JsonResponse({"success": True, "notification": _serialize_notification(notification)})


@login_required
@require_http_methods(["POST"])
@csrf_protect
def notification_mark_all_read(request):
    """Mark all unread notifications as read for the current user."""
    Notification.objects.filter(recipient=request.user, is_read=False).update(is_read=True, updated_at=timezone.now())
    return JsonResponse({"success": True})


@login_required
@require_http_methods(["GET", "POST"])
@csrf_protect
def notification_settings_api(request):
    """Read or update the current user's notification preferences."""
    ensure_notification_settings(request.user)
    if request.method == "POST":
        payload = json.loads(request.body or "{}")
        category = payload.get("category", "").strip()
        setting = get_object_or_404(NotificationSetting, user=request.user, category=category)
        for field in ("in_app_enabled", "email_enabled", "sms_enabled", "whatsapp_enabled"):
            if field in payload:
                setattr(setting, field, bool(payload.get(field)))
        setting.save()

    settings_payload = [
        {
            "category": item.category,
            "label": item.get_category_display(),
            "in_app_enabled": item.in_app_enabled,
            "email_enabled": item.email_enabled,
            "sms_enabled": item.sms_enabled,
            "whatsapp_enabled": item.whatsapp_enabled,
        }
        for item in NotificationSetting.objects.filter(user=request.user).order_by("category")
    ]
    return JsonResponse({"success": True, "items": settings_payload})


@login_required
def communications_page(request):
    """Render the communication hub and optionally preload a thread."""
    selected_thread_id = request.GET.get("thread", "").strip()
    threads = list(get_recent_threads_for_user(request.user, limit=18))
    selected_thread = None
    if selected_thread_id:
        selected_thread = get_object_or_404(CommunicationThread, id=selected_thread_id)
        if not can_access_thread(request.user, selected_thread):
            selected_thread = None

    return render(
        request,
        "core/communications_center.html",
        {
            "active_page": "communications",
            "communication_threads": threads,
            "selected_thread": selected_thread,
            "communication_scope": request.GET.get("scope", ""),
        },
    )


@login_required
def communication_threads_api(request):
    """Return filtered collaboration threads visible to the current user."""
    scope = request.GET.get("scope", "").strip()
    status_filter = request.GET.get("status", "").strip()
    search_query = request.GET.get("q", "").strip().lower()
    target_scope = request.GET.get("target_scope", "").strip()
    object_id = request.GET.get("object_id", "").strip()

    threads = list(get_recent_threads_for_user(request.user, limit=40))
    if scope:
        threads = [thread for thread in threads if thread.scope == scope]
    if status_filter:
        threads = [thread for thread in threads if thread.status == status_filter]
    if target_scope and object_id:
        threads = [thread for thread in threads if thread.scope == target_scope and str(thread.object_id) == object_id]
    if search_query:
        threads = [
            thread
            for thread in threads
            if search_query in thread.title.lower()
            or any(search_query in str(tag).lower() for tag in (thread.tags or []))
        ]
    return JsonResponse({"success": True, "items": [_serialize_thread(thread, request.user) for thread in threads]})


@login_required
@require_http_methods(["POST"])
@csrf_protect
def communication_thread_create(request):
    """Create a new collaboration thread and its initial comment."""
    payload = json.loads(request.body or "{}")
    title = payload.get("title", "").strip()
    scope = payload.get("scope", "").strip()
    body = payload.get("body", "").strip()
    if not title or not scope or not body:
        return JsonResponse({"success": False, "error": "Title, scope, and comment body are required."}, status=400)

    content_object = _resolve_thread_target(scope, payload.get("object_id"))
    assigned_to = User.objects.filter(username=payload.get("assigned_to", "").strip()).first()
    visibility = payload.get("visibility") or []
    tags = payload.get("tags") or []

    thread = create_thread_for_object(
        title=title,
        scope=scope,
        created_by=request.user,
        content_object=content_object,
        assigned_to=assigned_to,
        tags=tags,
        visibility=visibility,
    )
    add_comment_entry(thread=thread, author=request.user, body=body, status_update=payload.get("status_update", "none"))
    return JsonResponse({"success": True, "thread": _serialize_thread(thread, request.user)})


@login_required
def communication_thread_detail(request, thread_id):
    """Return a thread detail payload and mark it as read for the user."""
    thread = get_object_or_404(CommunicationThread.objects.select_related("created_by", "assigned_to"), id=thread_id)
    if not can_access_thread(request.user, thread):
        return JsonResponse({"success": False, "error": "You do not have access to this thread."}, status=403)

    touch_thread_participant(thread, request.user, mark_read=True)
    entries = thread.entries.select_related("author", "parent").prefetch_related("mentions")
    return JsonResponse(
        {
            "success": True,
            "thread": _serialize_thread(thread, request.user),
            "entries": [_serialize_comment(entry) for entry in entries],
        }
    )


@login_required
@require_http_methods(["POST"])
@csrf_protect
def communication_thread_comment(request, thread_id):
    """Append a comment to an existing collaboration thread."""
    thread = get_object_or_404(CommunicationThread, id=thread_id)
    if not can_access_thread(request.user, thread):
        return JsonResponse({"success": False, "error": "You do not have access to this thread."}, status=403)

    body = request.POST.get("body", "").strip()
    if not body:
        return JsonResponse({"success": False, "error": "Comment body is required."}, status=400)
    parent_id = request.POST.get("parent_id", "").strip()
    parent = CommentEntry.objects.filter(thread=thread, id=parent_id).first() if parent_id else None
    entry = add_comment_entry(
        thread=thread,
        author=request.user,
        body=body,
        parent=parent,
        attachment=request.FILES.get("attachment"),
        status_update=request.POST.get("status_update", "none"),
    )
    touch_thread_participant(thread, request.user, mark_read=True)
    return JsonResponse({"success": True, "entry": _serialize_comment(entry)})


@login_required
@require_http_methods(["POST"])
@csrf_protect
def communication_thread_mark_read(request, thread_id):
    """Mark a collaboration thread as read for the current user."""
    thread = get_object_or_404(CommunicationThread, id=thread_id)
    if not can_access_thread(request.user, thread):
        return JsonResponse({"success": False, "error": "You do not have access to this thread."}, status=403)
    touch_thread_participant(thread, request.user, mark_read=True)
    return JsonResponse({"success": True})
