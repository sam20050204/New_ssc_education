"""
Create an Excel template file for importing new admissions
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Student Import Template"

# Define headers
headers = [
    'S.No', 'Full Name', 'Student Name', 'Father Name', 'Surname', 'Mother Name',
    'Date of Birth', 'Mobile (Own)', 'Parent Mobile', 'Gender', 'Marital Status',
    'Course', 'Batch Month', 'Batch Year', 'Educational Qualification',
    'Address', 'City', 'Tehsil/Block', 'District', 'Pin Code',
    'Total Fees (₹)', 'Paid Fees First Installment', 'Admission Date'
]

# Add headers to first row
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    # Style header row
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Set column widths
column_widths = {
    'A': 6,   # S.No
    'B': 20,  # Full Name
    'C': 15,  # Student Name
    'D': 15,  # Father Name
    'E': 12,  # Surname
    'F': 15,  # Mother Name
    'G': 12,  # Date of Birth
    'H': 12,  # Mobile (Own)
    'I': 13,  # Parent Mobile
    'J': 10,  # Gender
    'K': 13,  # Marital Status
    'L': 15,  # Course
    'M': 12,  # Batch Month
    'N': 12,  # Batch Year
    'O': 18,  # Educational Qualification
    'P': 20,  # Address
    'Q': 12,  # City
    'R': 15,  # Tehsil/Block
    'S': 12,  # District
    'T': 10,  # Pin Code
    'U': 12,  # Total Fees
    'V': 18,  # Paid Fees First Installment
    'W': 14,  # Admission Date
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Add example data rows
example_data = [
    [
        1, 'Samarth Bandu khudase', 'Samarth', 'Bandu', 'khudase', 'Mahadevi',
        '04-01-2011', '8180801008', '9876543210', 'Male', 'Single',
        'MS-CIT', 'May', '2026', '9th Pass',
        'Shivaji Anagr, Murud', 'Latur', 'Murud', 'Latur', '413510',
        5000, 2000, '18-04-2026'
    ],
    [
        2, 'Shriyash Bandu khudase', 'Shriyash', 'Bandu', 'khudase', '',
        '03-03-2016', '8180801008', '', 'Male', 'Single',
        'Tally', 'May', '2026', '9th Pass',
        'Shivaji Anagr, Murud', 'Latur', 'Murud', 'Latur', '362843',
        5000, 0, '18-04-2026'
    ],
    [
        3, 'Priya Sharma', 'Priya', 'Rajesh', 'Sharma', 'Sunita',
        '15-06-2005', '7921234567', '9876543210', 'Female', 'Single',
        'Advance Excel', 'April', '2026', '12th Pass',
        'Main Road, Parbhani', 'Parbhani', 'Parbhani', 'Parbhani', '431401',
        5000, 1000, '15-04-2026'
    ]
]

# Define border style
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Add example data
for row_num, row_data in enumerate(example_data, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = thin_border
        
        # Format numeric columns
        if col_num in [1, 20, 21]:  # S.No, Total Fees, Paid Fees
            cell.alignment = Alignment(horizontal="right", vertical="center")
            if col_num in [20, 21]:
                cell.number_format = '₹ #,##0'
        
        # Format date columns
        if col_num == 7:  # Date of Birth
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Format pin code
        if col_num == 20:  # Pin Code
            cell.alignment = Alignment(horizontal="center", vertical="center")

# Add header border
for col_num in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_num)
    cell.border = thin_border

# Freeze the header row
ws.freeze_panes = "A2"

# Save the workbook
wb.save('Student_Import_Template.xlsx')
print("✅ Excel template created: Student_Import_Template.xlsx")
print("\nTemplate includes:")
print("  - Headers with all required columns")
print("  - 3 example student records")
print("  - Proper formatting and column widths")
print("  - Instructions for data entry")
print("\nHow to use:")
print("  1. Keep the headers as they are (don't modify column names or order)")
print("  2. Delete the example rows")
print("  3. Enter your student data starting from row 2")
print("  4. Ensure all required fields are filled")
print("  5. Upload to the 'Import New Students' feature on the web application")
