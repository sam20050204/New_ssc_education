import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.http import require_http_methods

from core.constants import BATCH_STATUS_CHOICES, TIME_SLOT_VALUES
from core.models import AdmittedStudent, Batch
from core.permissions import ROLE_ADMIN, ROLE_SUPER_ADMIN, roles_required
from core.services.batch_service import (
    ENDED_BATCH_STATUSES,
    archive_batch,
    create_batch,
    get_active_batch_summaries,
    get_batch_dashboard_data,
    get_batch_report_data,
    get_batch_transition_preview,
    get_batch_transition_students,
    get_ended_batch_summaries,
    update_batch_lifecycle,
)


def _common_batch_filters():
    months = (
        AdmittedStudent.objects.exclude(batch_month__isnull=True)
        .exclude(batch_month="")
        .values_list("batch_month", flat=True)
        .distinct()
        .order_by("batch_month")
    )
    years = (
        AdmittedStudent.objects.exclude(batch_year__isnull=True)
        .exclude(batch_year="")
        .values_list("batch_year", flat=True)
        .distinct()
        .order_by("-batch_year")
    )
    courses = (
        AdmittedStudent.objects.exclude(course__isnull=True)
        .exclude(course="")
        .values_list("course", flat=True)
        .distinct()
        .order_by("course")
    )
    return {"filter_months": months, "filter_years": years, "filter_courses": courses, "status_choices": BATCH_STATUS_CHOICES}


@login_required
@roles_required("Super Admin", "Admin", "Attendance Manager")
def batch_overview_dashboard(request):
    context = get_batch_dashboard_data()
    context["active_page"] = "batch_overview"
    return render(request, "core/timetable/batch_overview.html", context)


@login_required
@roles_required("Super Admin", "Admin", "Counselor", "Attendance Manager")
def active_batches(request):
    filters = {
        "month": request.GET.get("month", "").strip(),
        "year": request.GET.get("year", "").strip(),
        "course": request.GET.get("course", "").strip(),
        "search": request.GET.get("search", "").strip(),
    }
    context = _common_batch_filters()
    context.update(
        {
            "batch_summaries": get_active_batch_summaries(**filters),
            "filters": filters,
            "active_page": "active_batches",
            "page_title": "Active Batches",
        }
    )
    return render(request, "core/batch_management/active_batches.html", context)


@login_required
@roles_required(ROLE_SUPER_ADMIN, ROLE_ADMIN)
def end_batch(request):
    selected_month = request.GET.get("batch_month", "").strip()
    selected_year = request.GET.get("batch_year", "").strip()
    selected_course = request.GET.get("course", "").strip()
    preview = None
    preview_students = []
    pending_fee_students = []
    has_filters = any([selected_course, selected_year, selected_month])
    active_batches = AdmittedStudent.objects.filter(is_archived=False, batch_status="active")
    active_batch_summaries = (
        get_active_batch_summaries(month=selected_month, year=selected_year, course=selected_course)
        if has_filters
        else []
    )

    if selected_month and selected_year and selected_course:
        try:
            preview = get_batch_transition_preview(
                batch_month=selected_month,
                batch_year=selected_year,
                course=selected_course,
                statuses=("active",),
            )
            preview_students = list(
                get_batch_transition_students(
                    batch_month=selected_month,
                    batch_year=selected_year,
                    course=selected_course,
                    statuses=("active",),
                )
            )
            pending_fee_students = [student for student in preview_students if student.remaining_fees > 0]
        except ValueError as exc:
            messages.error(request, str(exc))

    context = _common_batch_filters()
    context["filter_months"] = (
        active_batches.exclude(batch_month__isnull=True)
        .exclude(batch_month="")
        .values_list("batch_month", flat=True)
        .distinct()
        .order_by("batch_month")
    )
    context["filter_years"] = (
        active_batches.exclude(batch_year__isnull=True)
        .exclude(batch_year="")
        .values_list("batch_year", flat=True)
        .distinct()
        .order_by("-batch_year")
    )
    context["filter_courses"] = (
        active_batches.exclude(course__isnull=True)
        .exclude(course="")
        .values_list("course", flat=True)
        .distinct()
        .order_by("course")
    )
    context.update(
        {
            "selected_month": selected_month,
            "selected_year": selected_year,
            "selected_course": selected_course,
            "active_batch_summaries": active_batch_summaries,
            "has_filters": has_filters,
            "preview": preview,
            "preview_students": preview_students,
            "pending_fee_students": pending_fee_students,
            "active_page": "end_batch",
            "page_title": "End Batch",
        }
    )
    return render(request, "core/batch_management/end_batch.html", context)


@login_required
@roles_required(ROLE_SUPER_ADMIN, ROLE_ADMIN)
@require_http_methods(["POST"])
@csrf_protect
def end_batch_confirm(request):
    batch_month = request.POST.get("batch_month", "").strip()
    batch_year = request.POST.get("batch_year", "").strip()
    course = request.POST.get("course", "").strip()
    remarks = request.POST.get("remarks", "").strip()
    override = request.POST.get("admin_override") == "on"

    try:
        affected = update_batch_lifecycle(
            batch_month=batch_month,
            batch_year=batch_year,
            course=course,
            action="end",
            actor=request.user,
            request=request,
            remarks=remarks,
            override=override,
        )
        messages.success(request, f"Batch {batch_month} {batch_year} ({course}) ended successfully for {affected} students.")
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect(f"{reverse('end_batch')}?batch_month={batch_month}&batch_year={batch_year}&course={course}")

    return redirect("ended_batches")


@login_required
@roles_required("Super Admin", "Admin", "Attendance Manager", "Counselor")
def ended_batches(request):
    filters = {
        "month": request.GET.get("month", "").strip(),
        "year": request.GET.get("year", "").strip(),
        "course": request.GET.get("course", "").strip(),
        "search": request.GET.get("search", "").strip(),
    }
    selected_batch_students = []
    pending_fee_students = []
    if filters["month"] and filters["year"] and filters["course"]:
        selected_batch_students = list(
            get_batch_transition_students(
                batch_month=filters["month"],
                batch_year=filters["year"],
                course=filters["course"],
                statuses=ENDED_BATCH_STATUSES,
            )
        )
        pending_fee_students = [student for student in selected_batch_students if student.remaining_fees > 0]

    context = _common_batch_filters()
    context.update(
        {
            "batch_summaries": get_ended_batch_summaries(**filters),
            "filters": filters,
            "selected_batch_students": selected_batch_students,
            "pending_fee_students": pending_fee_students,
            "active_page": "ended_batches",
            "page_title": "Ended Batches",
        }
    )
    return render(request, "core/batch_management/ended_batches.html", context)


@login_required
@roles_required(ROLE_SUPER_ADMIN, ROLE_ADMIN)
def restore_batch(request):
    filters = {
        "month": request.GET.get("month", "").strip(),
        "year": request.GET.get("year", "").strip(),
        "course": request.GET.get("course", "").strip(),
        "search": request.GET.get("search", "").strip(),
    }
    context = _common_batch_filters()
    context.update(
        {
            "batch_summaries": get_ended_batch_summaries(**filters),
            "filters": filters,
            "active_page": "restore_batch",
            "page_title": "Restore Batch",
        }
    )
    return render(request, "core/batch_management/restore_batch.html", context)


@login_required
@roles_required(ROLE_SUPER_ADMIN, ROLE_ADMIN)
@require_http_methods(["POST"])
@csrf_protect
def restore_batch_confirm(request):
    batch_month = request.POST.get("batch_month", "").strip()
    batch_year = request.POST.get("batch_year", "").strip()
    course = request.POST.get("course", "").strip()
    remarks = request.POST.get("remarks", "").strip()

    try:
        affected = update_batch_lifecycle(
            batch_month=batch_month,
            batch_year=batch_year,
            course=course,
            action="restore",
            actor=request.user,
            request=request,
            remarks=remarks,
            override=True,
        )
        messages.success(request, f"Batch {batch_month} {batch_year} ({course}) restored successfully for {affected} students.")
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect("restore_batch")

    return redirect("active_batches")


@login_required
@roles_required("Super Admin", "Admin", "Attendance Manager", "Counselor", "Accountant")
def batch_reports(request):
    filters = {
        "month": request.GET.get("month", "").strip(),
        "year": request.GET.get("year", "").strip(),
        "course": request.GET.get("course", "").strip(),
        "status": request.GET.get("status", "").strip(),
    }
    selected_batch_students = []
    pending_fee_students = []
    if filters["month"] and filters["year"] and filters["course"]:
        selected_statuses = (filters["status"],) if filters["status"] else None
        selected_batch_students = list(
            get_batch_transition_students(
                batch_month=filters["month"],
                batch_year=filters["year"],
                course=filters["course"],
                statuses=selected_statuses,
            )
        )
        pending_fee_students = [student for student in selected_batch_students if student.remaining_fees > 0]

    context = _common_batch_filters()
    context.update(
        {
            "report_data": get_batch_report_data(**filters),
            "filters": filters,
            "selected_batch_students": selected_batch_students,
            "pending_fee_students": pending_fee_students,
            "active_page": "batch_reports",
            "page_title": "Batch Reports",
        }
    )
    return render(request, "core/batch_management/batch_reports.html", context)


@login_required
@roles_required("Super Admin", "Admin", "Attendance Manager", "Counselor", "Accountant")
def export_batch_reports(request):
    filters = {
        "month": request.GET.get("month", "").strip(),
        "year": request.GET.get("year", "").strip(),
        "course": request.GET.get("course", "").strip(),
        "status": request.GET.get("status", "").strip(),
    }
    report_data = get_batch_report_data(**filters)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Batch Reports"
    headers = ["Batch Month", "Batch Year", "Status", "Students", "Courses", "Paid Fees", "Pending Fees"]
    sheet.append(headers)
    header_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in report_data["summaries"]:
        sheet.append(
            [
                row["batch_month"],
                row["batch_year"],
                row["batch_status"],
                row["student_count"],
                row["course_names"],
                float(row["paid_fees"]),
                float(row["pending_fees"]),
            ]
        )

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="batch_reports.xlsx"'
    workbook.save(response)
    return response


@login_required
@roles_required("Super Admin", "Admin", "Attendance Manager")
@require_http_methods(["POST"])
@csrf_protect
def create_batch_view(request):
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Invalid JSON payload."}, status=400)

    batch_type = data.get("batch_type", "").strip()
    time_slot = data.get("time_slot", "").strip()
    capacity = int(data.get("capacity") or 50)

    if batch_type not in {"Theory", "Practical"}:
        return JsonResponse({"success": False, "error": "Invalid batch type."}, status=400)
    if time_slot not in TIME_SLOT_VALUES:
        return JsonResponse({"success": False, "error": "Invalid time slot."}, status=400)
    if Batch.objects.filter(batch_type=batch_type, time_slot=time_slot, course__isnull=True, is_archived=False).exists():
        return JsonResponse({"success": False, "error": "Batch already exists for that slot."}, status=400)

    batch = create_batch(batch_type=batch_type, time_slot=time_slot, capacity=max(1, capacity), actor=request.user, request=request)
    return JsonResponse({"success": True, "batch": {"id": batch.id, "type": batch.batch_type, "time_slot": batch.time_slot, "capacity": batch.capacity}})


@login_required
@roles_required("Super Admin", "Admin", "Attendance Manager")
@require_http_methods(["POST"])
@csrf_protect
def delete_batch(request, batch_id):
    batch = get_object_or_404(Batch, pk=batch_id, is_archived=False)
    archive_batch(batch=batch, actor=request.user, request=request)
    return JsonResponse({"success": True, "message": "Batch archived successfully."})
